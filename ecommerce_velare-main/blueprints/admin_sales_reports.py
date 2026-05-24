from flask import Blueprint, render_template, make_response, request
from database.db_config import get_supabase_client
import csv
from io import StringIO, BytesIO
from datetime import datetime, timedelta
from collections import defaultdict
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.platypus import (
    SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak,
    KeepTogether
)
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch, mm
from reportlab.lib.enums import TA_CENTER, TA_RIGHT, TA_LEFT
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

admin_sales_reports_bp = Blueprint('admin_sales_reports', __name__)

# ============================================================================
# DASHBOARD ROUTE (page render)
# ============================================================================

@admin_sales_reports_bp.route('/admin/sales-reports')
def admin_sales_reports():
    print("📊 Loading admin sales reports...")
    supabase = get_supabase_client()

    try:
        # Get total sales from delivered orders
        orders_response = supabase.table('orders').select(
            'order_id, total_amount, commission_amount, seller_id'
        ).eq('order_status', 'delivered').execute()
        total_sales = sum(order['total_amount'] for order in orders_response.data if order.get('total_amount'))

        # Get total delivered orders
        total_orders = len(orders_response.data)

        # Get total commission from delivered orders
        total_commission = sum(order['commission_amount'] for order in orders_response.data if order.get('commission_amount'))

        # Get sales by category - Show ALL predefined categories even with 0 sales
        predefined_categories = [
            'Dresses', 'Skirts', 'Tops', 'Blouses', 'Activewear',
            'Yoga Pants', 'Lingerie', 'Sleepwear', 'Jackets', 'Coats',
            'Shoes', 'Accessories'
        ]

        # Fetch all order items with product info for delivered orders
        order_items_response = supabase.table('order_items').select(
            'order_id, product_id, subtotal, quantity'
        ).execute()
        products_response = supabase.table('products').select(
            'product_id, category'
        ).execute()

        # Create product category map
        product_categories = {p['product_id']: p['category'] for p in products_response.data}

        # Get delivered order IDs
        delivered_order_ids = set(order['order_id'] for order in orders_response.data)

        # Group by category
        category_stats = {cat: {'total_sales': 0.0, 'order_count': 0, 'items_sold': 0} for cat in predefined_categories}
        order_categories = {}

        for item in order_items_response.data:
            if item['order_id'] in delivered_order_ids:
                category = product_categories.get(item['product_id'])
                if category in category_stats:
                    category_stats[category]['total_sales'] += float(item['subtotal']) if item.get('subtotal') else 0
                    category_stats[category]['items_sold'] += item['quantity']

                    if category not in order_categories:
                        order_categories[category] = set()
                    order_categories[category].add(item['order_id'])

        for cat in category_stats:
            if cat in order_categories:
                category_stats[cat]['order_count'] = len(order_categories[cat])

        formatted_categories = []
        for cat in predefined_categories:
            stats = category_stats[cat]
            formatted_categories.append({
                'category': cat,
                'total_sales': stats['total_sales'],
                'order_count': stats['order_count'],
                'items_sold': stats['items_sold'],
                'percentage': (stats['total_sales'] / total_sales * 100) if total_sales > 0 else 0
            })
        formatted_categories.sort(key=lambda x: x['total_sales'], reverse=True)

        # Top sellers
        sellers_response = supabase.table('sellers').select(
            'seller_id, shop_name, first_name, last_name'
        ).execute()
        sellers_map = {s['seller_id']: s for s in sellers_response.data}

        seller_stats = {}
        for order in orders_response.data:
            seller_id = order['seller_id']
            if seller_id not in seller_stats:
                seller_stats[seller_id] = {'total_sales': 0.0, 'order_count': 0}
            seller_stats[seller_id]['total_sales'] += float(order['total_amount'])
            seller_stats[seller_id]['order_count'] += 1

        formatted_sellers = []
        for seller_id, sstats in seller_stats.items():
            if seller_id in sellers_map:
                seller = sellers_map[seller_id]
                formatted_sellers.append({
                    'seller_id': seller_id,
                    'store_name': seller['shop_name'],
                    'name': f"{seller['first_name']} {seller['last_name']}",
                    'total_sales': sstats['total_sales'],
                    'order_count': sstats['order_count'],
                    'avg_order_value': sstats['total_sales'] / sstats['order_count'] if sstats['order_count'] > 0 else 0
                })
        formatted_sellers.sort(key=lambda x: x['total_sales'], reverse=True)
        formatted_sellers = formatted_sellers[:10]

        # Sales trend (last 7 days)
        seven_days_ago = (datetime.utcnow() - timedelta(days=7)).isoformat()
        trend_response = supabase.table('orders').select(
            'created_at, total_amount, order_id'
        ).eq('order_status', 'delivered').gte('created_at', seven_days_ago).execute()

        daily_stats = defaultdict(lambda: {'sales': 0.0, 'orders': 0})
        for order in trend_response.data:
            created_at = datetime.fromisoformat(order['created_at'].replace('Z', '+00:00'))
            date_key = created_at.strftime('%Y-%m-%d')
            daily_stats[date_key]['sales'] += float(order['total_amount'])
            daily_stats[date_key]['orders'] += 1

        formatted_trend = []
        for date_key in sorted(daily_stats.keys()):
            tstats = daily_stats[date_key]
            date_obj = datetime.strptime(date_key, '%Y-%m-%d')
            formatted_trend.append({
                'date': date_obj.strftime('%b %d'),
                'sales': tstats['sales'],
                'orders': tstats['orders']
            })

        # Detailed sales (last 20 delivered orders)
        detailed_orders = supabase.table('orders').select(
            'order_id, order_number, total_amount, commission_amount, created_at, buyer_id, seller_id'
        ).eq('order_status', 'delivered').order('created_at', desc=True).limit(20).execute()

        buyer_ids = list(set(order['buyer_id'] for order in detailed_orders.data))
        seller_ids = list(set(order['seller_id'] for order in detailed_orders.data))

        buyers_response = supabase.table('buyers').select(
            'buyer_id, first_name, last_name'
        ).in_('buyer_id', buyer_ids).execute() if buyer_ids else type('R', (), {'data': []})()
        sellers_detail_response = supabase.table('sellers').select(
            'seller_id, shop_name'
        ).in_('seller_id', seller_ids).execute() if seller_ids else type('R', (), {'data': []})()

        buyers_map = {b['buyer_id']: b for b in buyers_response.data}
        sellers_detail_map = {s['seller_id']: s for s in sellers_detail_response.data}

        order_ids = [order['order_id'] for order in detailed_orders.data]
        items_response = supabase.table('order_items').select(
            'order_id, product_id'
        ).in_('order_id', order_ids).execute() if order_ids else type('R', (), {'data': []})()

        order_items_map = defaultdict(list)
        for item in items_response.data:
            order_items_map[item['order_id']].append(item['product_id'])

        formatted_detailed = []
        for order in detailed_orders.data:
            buyer = buyers_map.get(order['buyer_id'], {})
            seller = sellers_detail_map.get(order['seller_id'], {})
            product_ids = order_items_map.get(order['order_id'], [])

            categories = set()
            for pid in product_ids:
                if pid in product_categories:
                    categories.add(product_categories[pid])

            created_at = datetime.fromisoformat(order['created_at'].replace('Z', '+00:00'))

            formatted_detailed.append({
                'order_id': order['order_id'],
                'order_number': order['order_number'],
                'amount': float(order['total_amount']),
                'commission': float(order['commission_amount']),
                'date': created_at.strftime('%b %d, %Y'),
                'buyer': f"{buyer.get('first_name', 'Unknown')} {buyer.get('last_name', '')}".strip(),
                'seller': seller.get('shop_name', 'Unknown'),
                'item_count': len(product_ids),
                'categories': ', '.join(categories) if categories else 'N/A'
            })

        stats = {
            'total_sales': total_sales,
            'total_orders': total_orders,
            'total_commission': total_commission,
            'sales_by_category': formatted_categories,
            'top_sellers': formatted_sellers,
            'sales_trend': formatted_trend,
            'detailed_sales': formatted_detailed
        }

        print(f"✅ Sales reports loaded successfully")

    except Exception as e:
        print(f"❌ Error fetching sales reports: {e}")
        stats = {
            'total_sales': 0.0,
            'total_orders': 0,
            'total_commission': 0.0,
            'sales_by_category': [],
            'top_sellers': [],
            'sales_trend': [],
            'detailed_sales': []
        }

    return render_template('admin/admin_sales_reports.html', stats=stats)


# ============================================================================
# SHARED DATA-FETCH HELPER (used by CSV / Excel / PDF exports)
# ============================================================================

def _fetch_sales_export_data(start_date, end_date, category):
    """Fetch and shape the sales-report data set used by all three exports.

    Returns a dict with:
        - period_label: human-readable period
        - generated_at: datetime for the report header
        - filter_category: the category filter (or '' for all)
        - summary: aggregate metrics
        - rows: list of per-order dicts ready to write to a table
        - by_category: list of category aggregates for the period+filter
        - by_seller: list of seller aggregates for the period+filter
    """
    supabase = get_supabase_client()

    # --- Build orders query within optional date range ---
    query = supabase.table('orders').select(
        'order_id, order_number, created_at, total_amount, commission_amount, '
        'shipping_fee, subtotal, discount_amount, buyer_id, seller_id, order_status'
    ).eq('order_status', 'delivered')

    if start_date:
        start_dt = datetime.strptime(start_date, '%Y-%m-%d').isoformat()
        query = query.gte('created_at', start_dt)
    if end_date:
        end_dt = datetime.strptime(end_date, '%Y-%m-%d').replace(
            hour=23, minute=59, second=59
        ).isoformat()
        query = query.lte('created_at', end_dt)

    sales_data = query.order('created_at', desc=True).execute().data or []

    # --- Apply optional category filter ---
    if category:
        products_response = supabase.table('products').select(
            'product_id'
        ).eq('category', category).execute()
        category_product_ids = [p['product_id'] for p in (products_response.data or [])]

        if category_product_ids:
            items_for_cat = supabase.table('order_items').select(
                'order_id'
            ).in_('product_id', category_product_ids).execute()
            filtered_order_ids = set(
                item['order_id'] for item in (items_for_cat.data or [])
            )
            sales_data = [
                order for order in sales_data
                if order['order_id'] in filtered_order_ids
            ]
        else:
            sales_data = []

    # --- Batch fetch related buyer / seller / item / product data ---
    if sales_data:
        buyer_ids = list({order['buyer_id'] for order in sales_data})
        seller_ids = list({order['seller_id'] for order in sales_data})
        order_ids = [order['order_id'] for order in sales_data]

        buyers_response = supabase.table('buyers').select(
            'buyer_id, first_name, last_name'
        ).in_('buyer_id', buyer_ids).execute()
        sellers_response = supabase.table('sellers').select(
            'seller_id, shop_name'
        ).in_('seller_id', seller_ids).execute()
        items_response = supabase.table('order_items').select(
            'order_id, product_id, quantity, subtotal'
        ).in_('order_id', order_ids).execute()

        product_ids = list({
            item['product_id']
            for item in (items_response.data or [])
        })
        products_response = supabase.table('products').select(
            'product_id, product_name, category'
        ).in_('product_id', product_ids).execute() if product_ids else type('R', (), {'data': []})()

        buyers_map = {b['buyer_id']: b for b in (buyers_response.data or [])}
        sellers_map = {s['seller_id']: s for s in (sellers_response.data or [])}
        products_map = {p['product_id']: p for p in (products_response.data or [])}

        order_items_map = defaultdict(list)
        for item in (items_response.data or []):
            order_items_map[item['order_id']].append(item)
    else:
        buyers_map = {}
        sellers_map = {}
        products_map = {}
        order_items_map = {}

    # --- Shape per-order rows ---
    rows = []
    by_seller_acc = defaultdict(lambda: {'sales': 0.0, 'commission': 0.0, 'orders': 0, 'items': 0})
    by_category_acc = defaultdict(lambda: {'sales': 0.0, 'orders': set(), 'items': 0})

    for order in sales_data:
        created_at = datetime.fromisoformat(order['created_at'].replace('Z', '+00:00'))

        buyer = buyers_map.get(order['buyer_id'], {})
        seller = sellers_map.get(order['seller_id'], {})

        items = order_items_map.get(order['order_id'], [])
        item_count = sum(item.get('quantity', 0) for item in items)

        product_names = []
        order_categories = set()
        for item in items:
            product = products_map.get(item['product_id'], {})
            if product.get('product_name'):
                product_names.append(f"{product['product_name']} ({item['quantity']}x)")
            if product.get('category'):
                order_categories.add(product['category'])

        total_amount = float(order.get('total_amount') or 0)
        commission = float(order.get('commission_amount') or 0)
        shipping = float(order.get('shipping_fee') or 0)
        discount = float(order.get('discount_amount') or 0)
        subtotal = float(order.get('subtotal') or (total_amount - shipping + discount))
        net_to_seller = total_amount - commission

        seller_name = seller.get('shop_name', 'Unknown Shop')
        buyer_name = f"{buyer.get('first_name', 'Unknown')} {buyer.get('last_name', '')}".strip() or 'Unknown'

        rows.append({
            'order_number': order['order_number'],
            'date': created_at.strftime('%Y-%m-%d'),
            'time': created_at.strftime('%I:%M %p'),
            'datetime_obj': created_at,
            'seller_name': seller_name,
            'buyer_name': buyer_name,
            'item_count': item_count,
            'subtotal': subtotal,
            'shipping': shipping,
            'discount': discount,
            'total_amount': total_amount,
            'commission': commission,
            'net_to_seller': net_to_seller,
            'categories': sorted(order_categories),
            'products': product_names,
        })

        # Seller aggregates
        by_seller_acc[seller_name]['sales'] += total_amount
        by_seller_acc[seller_name]['commission'] += commission
        by_seller_acc[seller_name]['orders'] += 1
        by_seller_acc[seller_name]['items'] += item_count

        # Category aggregates (an order can belong to multiple categories)
        for cat in order_categories:
            by_category_acc[cat]['sales'] += total_amount / max(len(order_categories), 1)
            by_category_acc[cat]['orders'].add(order['order_id'])
            by_category_acc[cat]['items'] += item_count

    # --- Aggregates ---
    total_orders = len(rows)
    total_sales = sum(r['total_amount'] for r in rows)
    total_commission = sum(r['commission'] for r in rows)
    total_shipping = sum(r['shipping'] for r in rows)
    total_items = sum(r['item_count'] for r in rows)
    avg_order = (total_sales / total_orders) if total_orders else 0
    commission_rate = (total_commission / total_sales * 100) if total_sales else 0

    by_seller = sorted(
        [
            {
                'seller': name,
                'orders': data['orders'],
                'items': data['items'],
                'sales': data['sales'],
                'commission': data['commission'],
                'net_to_seller': data['sales'] - data['commission'],
            }
            for name, data in by_seller_acc.items()
        ],
        key=lambda x: x['sales'],
        reverse=True,
    )

    by_category = sorted(
        [
            {
                'category': cat,
                'orders': len(data['orders']),
                'items': data['items'],
                'sales': data['sales'],
                'percentage': (data['sales'] / total_sales * 100) if total_sales else 0,
            }
            for cat, data in by_category_acc.items()
        ],
        key=lambda x: x['sales'],
        reverse=True,
    )

    # --- Period label ---
    if start_date and end_date:
        try:
            start_label = datetime.strptime(start_date, '%Y-%m-%d').strftime('%B %d, %Y')
            end_label = datetime.strptime(end_date, '%Y-%m-%d').strftime('%B %d, %Y')
            period_label = f"{start_label} – {end_label}"
        except ValueError:
            period_label = f"{start_date} – {end_date}"
    elif start_date:
        period_label = f"From {start_date}"
    elif end_date:
        period_label = f"Up to {end_date}"
    else:
        period_label = "All Time"

    return {
        'period_label': period_label,
        'generated_at': datetime.now(),
        'filter_category': category or 'All Categories',
        'summary': {
            'total_orders': total_orders,
            'total_sales': total_sales,
            'total_commission': total_commission,
            'total_shipping': total_shipping,
            'total_items': total_items,
            'avg_order': avg_order,
            'commission_rate': commission_rate,
            'net_to_sellers': total_sales - total_commission,
        },
        'rows': rows,
        'by_seller': by_seller,
        'by_category': by_category,
    }


def _peso(value):
    """Format a number as Philippine Peso with thousand separators."""
    return f"P{value:,.2f}"


# ============================================================================
# CSV EXPORT
# ============================================================================

@admin_sales_reports_bp.route('/admin/sales-reports/export-csv')
def export_csv():
    """Export sales reports to CSV format."""
    print("📄 Exporting sales report to CSV...")

    try:
        start_date = request.args.get('start_date', '')
        end_date = request.args.get('end_date', '')
        category = request.args.get('category', '')

        data = _fetch_sales_export_data(start_date, end_date, category)
        summary = data['summary']

        output = StringIO()
        writer = csv.writer(output)

        # Report header
        writer.writerow(['VELARE E-COMMERCE PLATFORM'])
        writer.writerow(['SALES REPORT'])
        writer.writerow([])
        writer.writerow(['Generated', data['generated_at'].strftime('%B %d, %Y at %I:%M %p')])
        writer.writerow(['Report Period', data['period_label']])
        writer.writerow(['Category Filter', data['filter_category']])
        writer.writerow(['Currency', 'Philippine Peso (PHP)'])
        writer.writerow([])

        # Executive summary
        writer.writerow(['--- EXECUTIVE SUMMARY ---'])
        writer.writerow(['Metric', 'Value'])
        writer.writerow(['Total Delivered Orders', f"{summary['total_orders']:,}"])
        writer.writerow(['Total Items Sold', f"{summary['total_items']:,}"])
        writer.writerow(['Gross Sales Revenue', _peso(summary['total_sales'])])
        writer.writerow(['Total Shipping Fees', _peso(summary['total_shipping'])])
        writer.writerow(['Platform Commission', _peso(summary['total_commission'])])
        writer.writerow(['Net Payout to Sellers', _peso(summary['net_to_sellers'])])
        writer.writerow(['Average Order Value', _peso(summary['avg_order'])])
        writer.writerow(['Effective Commission Rate', f"{summary['commission_rate']:.2f}%"])
        writer.writerow([])

        # Sales by category
        writer.writerow(['--- SALES BY CATEGORY ---'])
        writer.writerow(['Category', 'Orders', 'Items Sold', 'Sales', '% of Total'])
        for c in data['by_category']:
            writer.writerow([
                c['category'], c['orders'], c['items'],
                _peso(c['sales']), f"{c['percentage']:.1f}%",
            ])
        writer.writerow([])

        # Sales by seller
        writer.writerow(['--- SALES BY SELLER ---'])
        writer.writerow(['Seller', 'Orders', 'Items', 'Sales', 'Commission', 'Net to Seller'])
        for s in data['by_seller']:
            writer.writerow([
                s['seller'], s['orders'], s['items'],
                _peso(s['sales']), _peso(s['commission']), _peso(s['net_to_seller']),
            ])
        writer.writerow([])

        # Detailed transactions
        writer.writerow(['--- DETAILED TRANSACTIONS ---'])
        writer.writerow([
            'Order No.', 'Date', 'Time', 'Seller', 'Buyer', 'Items',
            'Subtotal', 'Shipping', 'Discount', 'Total',
            'Commission', 'Net to Seller', 'Categories', 'Products',
        ])
        for r in data['rows']:
            writer.writerow([
                r['order_number'], r['date'], r['time'],
                r['seller_name'], r['buyer_name'], r['item_count'],
                _peso(r['subtotal']), _peso(r['shipping']), _peso(r['discount']),
                _peso(r['total_amount']), _peso(r['commission']), _peso(r['net_to_seller']),
                ', '.join(r['categories']) if r['categories'] else 'N/A',
                '; '.join(r['products']) if r['products'] else 'N/A',
            ])

        writer.writerow([])
        writer.writerow(['Total Records', summary['total_orders']])
        writer.writerow(['END OF REPORT'])

        output.seek(0)
        # Add UTF-8 BOM so Excel opens the peso symbol correctly even on Windows.
        body = '\ufeff' + output.getvalue()
        response = make_response(body)
        response.headers['Content-Type'] = 'text/csv; charset=utf-8'
        filename = f"velare_sales_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
        response.headers['Content-Disposition'] = f'attachment; filename={filename}'
        print(f"✅ CSV export completed ({summary['total_orders']} rows)")
        return response

    except Exception as e:
        print(f"❌ Error exporting CSV: {e}")
        import traceback
        traceback.print_exc()
        return f"Error exporting CSV: {str(e)}", 500


# ============================================================================
# EXCEL EXPORT (multi-sheet, styled)
# ============================================================================

# Reusable Excel style palette (Velare brand-friendly).
_XL_BRAND = "2C2236"           # deep brand color for cover title
_XL_ACCENT = "D3BD9B"          # accent for table headers
_XL_HEADER_TEXT = "FFFFFF"
_XL_BAND = "F8F4EC"            # alternating row band
_XL_BORDER = "C9BFAE"
_XL_SUMMARY_LABEL = "EFE7D6"


def _xl_thin_border():
    side = Side(border_style='thin', color=_XL_BORDER)
    return Border(left=side, right=side, top=side, bottom=side)


def _xl_write_header(ws, title, subtitle, generated, period, category, total_columns):
    """Write the report title block at the top of a worksheet."""
    last_col_letter = get_column_letter(total_columns)

    ws.merge_cells(f'A1:{last_col_letter}1')
    cell = ws['A1']
    cell.value = title
    cell.font = Font(name='Calibri', size=18, bold=True, color=_XL_BRAND)
    cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 28

    ws.merge_cells(f'A2:{last_col_letter}2')
    cell = ws['A2']
    cell.value = subtitle
    cell.font = Font(name='Calibri', size=12, italic=True, color="555555")
    cell.alignment = Alignment(horizontal='center', vertical='center')

    # Metadata block
    meta = [
        ('Generated', generated.strftime('%B %d, %Y at %I:%M %p')),
        ('Report Period', period),
        ('Category Filter', category),
        ('Currency', 'Philippine Peso (PHP)'),
    ]
    for offset, (label, value) in enumerate(meta):
        row = 4 + offset
        ws[f'A{row}'] = label
        ws[f'A{row}'].font = Font(bold=True, color=_XL_BRAND)
        ws[f'A{row}'].alignment = Alignment(horizontal='left', vertical='center')

        ws.merge_cells(f'B{row}:{last_col_letter}{row}')
        ws[f'B{row}'] = value
        ws[f'B{row}'].alignment = Alignment(horizontal='left', vertical='center')

    return 4 + len(meta) + 1  # next free row


def _xl_section_title(ws, row, total_columns, text):
    last_col_letter = get_column_letter(total_columns)
    ws.merge_cells(f'A{row}:{last_col_letter}{row}')
    cell = ws[f'A{row}']
    cell.value = text
    cell.font = Font(name='Calibri', size=13, bold=True, color="FFFFFF")
    cell.fill = PatternFill(start_color=_XL_BRAND, end_color=_XL_BRAND, fill_type='solid')
    cell.alignment = Alignment(horizontal='left', vertical='center', indent=1)
    ws.row_dimensions[row].height = 22
    return row + 1


def _xl_table_headers(ws, row, headers):
    border = _xl_thin_border()
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = Font(bold=True, color=_XL_HEADER_TEXT)
        cell.fill = PatternFill(start_color=_XL_ACCENT, end_color=_XL_ACCENT, fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = border
    ws.row_dimensions[row].height = 26
    return row + 1


def _xl_data_row(ws, row, values, alignments=None, number_formats=None, banded=False):
    border = _xl_thin_border()
    fill = PatternFill(start_color=_XL_BAND, end_color=_XL_BAND, fill_type='solid') if banded else None
    for col, value in enumerate(values, start=1):
        cell = ws.cell(row=row, column=col, value=value)
        cell.border = border
        if fill:
            cell.fill = fill
        if alignments and col - 1 < len(alignments):
            cell.alignment = alignments[col - 1]
        else:
            cell.alignment = Alignment(horizontal='left', vertical='center')
        if number_formats and col - 1 < len(number_formats) and number_formats[col - 1]:
            cell.number_format = number_formats[col - 1]


def _xl_set_widths(ws, widths):
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width


@admin_sales_reports_bp.route('/admin/sales-reports/export-excel')
def export_excel():
    """Export sales reports to a styled multi-sheet Excel workbook."""
    print("📊 Exporting sales report to Excel...")

    try:
        start_date = request.args.get('start_date', '')
        end_date = request.args.get('end_date', '')
        category = request.args.get('category', '')

        data = _fetch_sales_export_data(start_date, end_date, category)
        summary = data['summary']

        wb = Workbook()

        # Common alignments / number formats reused across sheets
        left = Alignment(horizontal='left', vertical='center', wrap_text=True)
        center = Alignment(horizontal='center', vertical='center')
        right = Alignment(horizontal='right', vertical='center')
        peso_fmt = '"P"#,##0.00'
        int_fmt = '#,##0'
        pct_fmt = '0.0"%"'

        # ----------------------------------------------------------------
        # Sheet 1: Summary
        # ----------------------------------------------------------------
        ws_summary = wb.active
        ws_summary.title = "Summary"
        next_row = _xl_write_header(
            ws_summary,
            "VELARE E-COMMERCE PLATFORM",
            "Sales Report — Executive Summary",
            data['generated_at'],
            data['period_label'],
            data['filter_category'],
            total_columns=2,
        )

        next_row = _xl_section_title(ws_summary, next_row, 2, "Key Metrics")
        next_row = _xl_table_headers(ws_summary, next_row, ["Metric", "Value"])

        metric_rows = [
            ("Total Delivered Orders", summary['total_orders'], int_fmt),
            ("Total Items Sold", summary['total_items'], int_fmt),
            ("Gross Sales Revenue", summary['total_sales'], peso_fmt),
            ("Total Shipping Fees", summary['total_shipping'], peso_fmt),
            ("Platform Commission", summary['total_commission'], peso_fmt),
            ("Net Payout to Sellers", summary['net_to_sellers'], peso_fmt),
            ("Average Order Value", summary['avg_order'], peso_fmt),
            ("Effective Commission Rate", summary['commission_rate'], pct_fmt),
        ]
        border = _xl_thin_border()
        for offset, (label, value, fmt) in enumerate(metric_rows):
            row = next_row + offset
            label_cell = ws_summary.cell(row=row, column=1, value=label)
            label_cell.font = Font(bold=True)
            label_cell.fill = PatternFill(start_color=_XL_SUMMARY_LABEL, end_color=_XL_SUMMARY_LABEL, fill_type='solid')
            label_cell.alignment = left
            label_cell.border = border

            value_cell = ws_summary.cell(row=row, column=2, value=value)
            value_cell.alignment = right
            value_cell.number_format = fmt
            value_cell.border = border
        next_row += len(metric_rows) + 1

        # Top sellers preview
        if data['by_seller']:
            next_row = _xl_section_title(ws_summary, next_row, 2, "Top 5 Sellers (by sales)")
            next_row = _xl_table_headers(ws_summary, next_row, ["Seller", "Sales"])
            for idx, s in enumerate(data['by_seller'][:5]):
                _xl_data_row(
                    ws_summary, next_row,
                    [s['seller'], s['sales']],
                    alignments=[left, right],
                    number_formats=[None, peso_fmt],
                    banded=(idx % 2 == 1),
                )
                next_row += 1
            next_row += 1

        _xl_set_widths(ws_summary, [38, 28])

        # ----------------------------------------------------------------
        # Sheet 2: Sales by Category
        # ----------------------------------------------------------------
        ws_cat = wb.create_sheet("By Category")
        next_row = _xl_write_header(
            ws_cat,
            "VELARE E-COMMERCE PLATFORM",
            "Sales Breakdown by Category",
            data['generated_at'],
            data['period_label'],
            data['filter_category'],
            total_columns=5,
        )
        next_row = _xl_section_title(ws_cat, next_row, 5, "Categories")
        next_row = _xl_table_headers(ws_cat, next_row, ["Category", "Orders", "Items Sold", "Sales", "% of Total"])

        if not data['by_category']:
            ws_cat.merge_cells(start_row=next_row, start_column=1, end_row=next_row, end_column=5)
            ws_cat.cell(row=next_row, column=1, value='No data for this period.').alignment = center
            next_row += 1
        else:
            for idx, c in enumerate(data['by_category']):
                _xl_data_row(
                    ws_cat, next_row,
                    [c['category'], c['orders'], c['items'], c['sales'], c['percentage']],
                    alignments=[left, center, center, right, right],
                    number_formats=[None, int_fmt, int_fmt, peso_fmt, pct_fmt],
                    banded=(idx % 2 == 1),
                )
                next_row += 1

        ws_cat.freeze_panes = ws_cat[f'A{next_row - len(data["by_category"]) if data["by_category"] else next_row}']
        _xl_set_widths(ws_cat, [22, 12, 14, 18, 14])

        # ----------------------------------------------------------------
        # Sheet 3: Sales by Seller
        # ----------------------------------------------------------------
        ws_sellers = wb.create_sheet("By Seller")
        next_row = _xl_write_header(
            ws_sellers,
            "VELARE E-COMMERCE PLATFORM",
            "Sales Breakdown by Seller",
            data['generated_at'],
            data['period_label'],
            data['filter_category'],
            total_columns=6,
        )
        next_row = _xl_section_title(ws_sellers, next_row, 6, "Sellers")
        next_row = _xl_table_headers(
            ws_sellers, next_row,
            ["Seller", "Orders", "Items", "Sales", "Commission", "Net to Seller"],
        )

        if not data['by_seller']:
            ws_sellers.merge_cells(start_row=next_row, start_column=1, end_row=next_row, end_column=6)
            ws_sellers.cell(row=next_row, column=1, value='No data for this period.').alignment = center
            next_row += 1
        else:
            for idx, s in enumerate(data['by_seller']):
                _xl_data_row(
                    ws_sellers, next_row,
                    [s['seller'], s['orders'], s['items'], s['sales'], s['commission'], s['net_to_seller']],
                    alignments=[left, center, center, right, right, right],
                    number_formats=[None, int_fmt, int_fmt, peso_fmt, peso_fmt, peso_fmt],
                    banded=(idx % 2 == 1),
                )
                next_row += 1

        _xl_set_widths(ws_sellers, [30, 10, 10, 18, 18, 20])

        # ----------------------------------------------------------------
        # Sheet 4: Transactions (the main detailed log)
        # ----------------------------------------------------------------
        ws_tx = wb.create_sheet("Transactions")
        next_row = _xl_write_header(
            ws_tx,
            "VELARE E-COMMERCE PLATFORM",
            "Detailed Transactions",
            data['generated_at'],
            data['period_label'],
            data['filter_category'],
            total_columns=14,
        )
        next_row = _xl_section_title(ws_tx, next_row, 14, "Delivered Orders")
        tx_headers = [
            "Order No.", "Date", "Time", "Seller", "Buyer", "Items",
            "Subtotal", "Shipping", "Discount", "Total",
            "Commission", "Net to Seller", "Categories", "Products",
        ]
        next_row = _xl_table_headers(ws_tx, next_row, tx_headers)
        first_data_row = next_row

        if not data['rows']:
            ws_tx.merge_cells(start_row=next_row, start_column=1, end_row=next_row, end_column=14)
            ws_tx.cell(row=next_row, column=1, value='No transactions for this period.').alignment = center
            next_row += 1
        else:
            tx_alignments = [
                left, center, center, left, left, center,
                right, right, right, right,
                right, right, left, left,
            ]
            tx_formats = [
                None, None, None, None, None, int_fmt,
                peso_fmt, peso_fmt, peso_fmt, peso_fmt,
                peso_fmt, peso_fmt, None, None,
            ]
            for idx, r in enumerate(data['rows']):
                _xl_data_row(
                    ws_tx, next_row,
                    [
                        r['order_number'], r['date'], r['time'],
                        r['seller_name'], r['buyer_name'], r['item_count'],
                        r['subtotal'], r['shipping'], r['discount'], r['total_amount'],
                        r['commission'], r['net_to_seller'],
                        ', '.join(r['categories']) if r['categories'] else 'N/A',
                        '; '.join(r['products']) if r['products'] else 'N/A',
                    ],
                    alignments=tx_alignments,
                    number_formats=tx_formats,
                    banded=(idx % 2 == 1),
                )
                next_row += 1

            # Totals row
            border = _xl_thin_border()
            totals = [
                "TOTAL", "", "", "", "", summary['total_items'],
                sum(r['subtotal'] for r in data['rows']),
                summary['total_shipping'],
                sum(r['discount'] for r in data['rows']),
                summary['total_sales'],
                summary['total_commission'],
                summary['net_to_sellers'],
                "", "",
            ]
            for col, val in enumerate(totals, start=1):
                cell = ws_tx.cell(row=next_row, column=col, value=val)
                cell.font = Font(bold=True, color=_XL_HEADER_TEXT)
                cell.fill = PatternFill(start_color=_XL_BRAND, end_color=_XL_BRAND, fill_type='solid')
                cell.border = border
                if col >= 7 and col <= 12 and col != 13:
                    cell.number_format = peso_fmt
                    cell.alignment = right
                elif col == 6:
                    cell.number_format = int_fmt
                    cell.alignment = center
                else:
                    cell.alignment = center
            next_row += 1

        # Freeze the header row
        ws_tx.freeze_panes = ws_tx[f'A{first_data_row}']
        _xl_set_widths(
            ws_tx,
            [14, 12, 10, 28, 24, 8, 14, 14, 14, 14, 14, 16, 22, 50],
        )

        # Save to memory
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        filename = f"velare_sales_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        response = make_response(buffer.getvalue())
        response.headers['Content-Type'] = (
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response.headers['Content-Disposition'] = f'attachment; filename={filename}'
        print(f"✅ Excel export completed ({summary['total_orders']} rows, 4 sheets)")
        return response

    except Exception as e:
        print(f"❌ Error exporting Excel: {e}")
        import traceback
        traceback.print_exc()
        return f"Error exporting Excel: {str(e)}", 500


# ============================================================================
# PDF EXPORT (landscape, sectioned, branded)
# ============================================================================

# Brand colors as ReportLab Color instances.
_PDF_BRAND = colors.HexColor("#2C2236")
_PDF_ACCENT = colors.HexColor("#D3BD9B")
_PDF_BAND = colors.HexColor("#F8F4EC")
_PDF_HEADER_TEXT = colors.white
_PDF_BORDER = colors.HexColor("#C9BFAE")
_PDF_MUTED = colors.HexColor("#6D6552")


def _pdf_styles():
    base = getSampleStyleSheet()
    return {
        'title': ParagraphStyle(
            name='VTitle', parent=base['Title'],
            fontName='Helvetica-Bold', fontSize=22, leading=26,
            alignment=TA_CENTER, textColor=_PDF_BRAND, spaceAfter=4,
        ),
        'subtitle': ParagraphStyle(
            name='VSubtitle', parent=base['Normal'],
            fontName='Helvetica-Oblique', fontSize=12, leading=14,
            alignment=TA_CENTER, textColor=_PDF_MUTED, spaceAfter=14,
        ),
        'meta': ParagraphStyle(
            name='VMeta', parent=base['Normal'],
            fontName='Helvetica', fontSize=9.5, leading=12,
            alignment=TA_LEFT, textColor=colors.HexColor('#333333'),
        ),
        'section': ParagraphStyle(
            name='VSection', parent=base['Heading2'],
            fontName='Helvetica-Bold', fontSize=13, leading=16,
            textColor=colors.white, backColor=_PDF_BRAND,
            leftIndent=8, rightIndent=8, spaceBefore=10, spaceAfter=6,
            borderPadding=(6, 8, 6, 8),
        ),
        'small': ParagraphStyle(
            name='VSmall', parent=base['Normal'],
            fontName='Helvetica', fontSize=8.5, leading=10,
            textColor=colors.HexColor('#333333'),
        ),
        'small_right': ParagraphStyle(
            name='VSmallRight', parent=base['Normal'],
            fontName='Helvetica', fontSize=8.5, leading=10,
            alignment=TA_RIGHT,
        ),
        'footer': ParagraphStyle(
            name='VFooter', parent=base['Normal'],
            fontName='Helvetica', fontSize=8, leading=10,
            textColor=_PDF_MUTED, alignment=TA_CENTER,
        ),
    }


def _pdf_make_table(headers, rows, col_widths, align_overrides=None):
    """Build a styled ReportLab Table.

    align_overrides: dict {col_index: 'LEFT'|'CENTER'|'RIGHT'} applied per column.
    """
    data = [headers] + rows if rows else [headers, ['(no data)'] + [''] * (len(headers) - 1)]
    tbl = Table(data, colWidths=col_widths, repeatRows=1)

    style = TableStyle([
        # Header
        ('BACKGROUND', (0, 0), (-1, 0), _PDF_ACCENT),
        ('TEXTCOLOR', (0, 0), (-1, 0), _PDF_HEADER_TEXT),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 9),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        ('VALIGN', (0, 0), (-1, 0), 'MIDDLE'),
        ('TOPPADDING', (0, 0), (-1, 0), 6),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 6),
        # Body
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 8.5),
        ('VALIGN', (0, 1), (-1, -1), 'MIDDLE'),
        ('TOPPADDING', (0, 1), (-1, -1), 4),
        ('BOTTOMPADDING', (0, 1), (-1, -1), 4),
        ('LEFTPADDING', (0, 0), (-1, -1), 4),
        ('RIGHTPADDING', (0, 0), (-1, -1), 4),
        # Borders
        ('GRID', (0, 0), (-1, -1), 0.4, _PDF_BORDER),
        # Banded rows
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, _PDF_BAND]),
    ])

    # Per-column alignment
    if align_overrides:
        for col_idx, alignment in align_overrides.items():
            style.add('ALIGN', (col_idx, 1), (col_idx, -1), alignment)

    tbl.setStyle(style)
    return tbl


def _pdf_summary_block(summary, period_label, category_label, generated):
    """Two-column summary panel with key metrics."""
    rows = [
        ['Total Delivered Orders', f"{summary['total_orders']:,}"],
        ['Total Items Sold', f"{summary['total_items']:,}"],
        ['Gross Sales Revenue', _peso(summary['total_sales'])],
        ['Total Shipping Fees', _peso(summary['total_shipping'])],
        ['Platform Commission', _peso(summary['total_commission'])],
        ['Net Payout to Sellers', _peso(summary['net_to_sellers'])],
        ['Average Order Value', _peso(summary['avg_order'])],
        ['Effective Commission Rate', f"{summary['commission_rate']:.2f}%"],
    ]

    tbl = Table(rows, colWidths=[2.5 * inch, 2.0 * inch])
    tbl.setStyle(TableStyle([
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('FONTNAME', (1, 0), (1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('TEXTCOLOR', (0, 0), (0, -1), _PDF_BRAND),
        ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('TOPPADDING', (0, 0), (-1, -1), 6),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ('LEFTPADDING', (0, 0), (-1, -1), 8),
        ('RIGHTPADDING', (0, 0), (-1, -1), 8),
        ('LINEBELOW', (0, 0), (-1, -2), 0.4, _PDF_BORDER),
        ('BACKGROUND', (0, 0), (0, -1), _PDF_BAND),
    ]))
    return tbl


def _on_page(canvas, doc, generated_at):
    """Footer drawn on every page."""
    canvas.saveState()
    canvas.setFont('Helvetica', 8)
    canvas.setFillColor(_PDF_MUTED)
    page_num = canvas.getPageNumber()
    footer_left = f"Velare Sales Report — Generated {generated_at.strftime('%B %d, %Y at %I:%M %p')}"
    footer_right = f"Page {page_num}"
    canvas.drawString(0.5 * inch, 0.4 * inch, footer_left)
    canvas.drawRightString(doc.pagesize[0] - 0.5 * inch, 0.4 * inch, footer_right)
    canvas.restoreState()


@admin_sales_reports_bp.route('/admin/sales-reports/export-pdf')
def export_pdf():
    """Export sales reports to a polished landscape PDF."""
    print("📄 Exporting sales report to PDF...")

    try:
        start_date = request.args.get('start_date', '')
        end_date = request.args.get('end_date', '')
        category = request.args.get('category', '')

        data = _fetch_sales_export_data(start_date, end_date, category)
        summary = data['summary']

        buffer = BytesIO()
        doc = SimpleDocTemplate(
            buffer,
            pagesize=landscape(A4),
            leftMargin=0.5 * inch, rightMargin=0.5 * inch,
            topMargin=0.5 * inch, bottomMargin=0.6 * inch,
            title="Velare Sales Report",
            author="Velare E-Commerce Platform",
        )

        styles = _pdf_styles()
        flow = []

        # ---- Header block ----
        flow.append(Paragraph("VELARE E-COMMERCE PLATFORM", styles['title']))
        flow.append(Paragraph("Sales Report", styles['subtitle']))

        meta_lines = [
            f"<b>Generated:</b> {data['generated_at'].strftime('%B %d, %Y at %I:%M %p')}",
            f"<b>Report Period:</b> {data['period_label']}",
            f"<b>Category Filter:</b> {data['filter_category']}",
            f"<b>Currency:</b> Philippine Peso (PHP)",
        ]
        for line in meta_lines:
            flow.append(Paragraph(line, styles['meta']))
        flow.append(Spacer(1, 12))

        # ---- Executive Summary ----
        flow.append(Paragraph("EXECUTIVE SUMMARY", styles['section']))
        flow.append(_pdf_summary_block(summary, data['period_label'], data['filter_category'], data['generated_at']))
        flow.append(Spacer(1, 16))

        # ---- Sales by Category ----
        flow.append(Paragraph("SALES BY CATEGORY", styles['section']))
        cat_rows = [
            [
                c['category'],
                f"{c['orders']:,}",
                f"{c['items']:,}",
                _peso(c['sales']),
                f"{c['percentage']:.1f}%",
            ]
            for c in data['by_category']
        ]
        flow.append(_pdf_make_table(
            headers=["Category", "Orders", "Items Sold", "Sales", "% of Total"],
            rows=cat_rows,
            col_widths=[2.4 * inch, 1.0 * inch, 1.2 * inch, 1.6 * inch, 1.2 * inch],
            align_overrides={0: 'LEFT', 1: 'CENTER', 2: 'CENTER', 3: 'RIGHT', 4: 'RIGHT'},
        ))
        flow.append(Spacer(1, 16))

        # ---- Top Sellers ----
        flow.append(Paragraph("TOP SELLERS BY SALES", styles['section']))
        seller_rows = [
            [
                s['seller'],
                f"{s['orders']:,}",
                f"{s['items']:,}",
                _peso(s['sales']),
                _peso(s['commission']),
                _peso(s['net_to_seller']),
            ]
            for s in data['by_seller'][:15]
        ]
        flow.append(_pdf_make_table(
            headers=["Seller", "Orders", "Items", "Sales", "Commission", "Net to Seller"],
            rows=seller_rows,
            col_widths=[3.0 * inch, 0.8 * inch, 0.8 * inch, 1.4 * inch, 1.4 * inch, 1.5 * inch],
            align_overrides={0: 'LEFT', 1: 'CENTER', 2: 'CENTER', 3: 'RIGHT', 4: 'RIGHT', 5: 'RIGHT'},
        ))

        # ---- Detailed Transactions ----
        flow.append(PageBreak())
        flow.append(Paragraph("DETAILED TRANSACTIONS", styles['section']))

        if not data['rows']:
            flow.append(Paragraph("No transactions for this period.", styles['meta']))
        else:
            tx_rows = []
            for r in data['rows']:
                seller_para = Paragraph(r['seller_name'], styles['small'])
                buyer_para = Paragraph(r['buyer_name'], styles['small'])
                cats = ', '.join(r['categories']) if r['categories'] else 'N/A'
                cat_para = Paragraph(cats, styles['small'])

                tx_rows.append([
                    r['order_number'],
                    f"{r['date']}\n{r['time']}",
                    seller_para,
                    buyer_para,
                    f"{r['item_count']:,}",
                    _peso(r['subtotal']),
                    _peso(r['shipping']),
                    _peso(r['total_amount']),
                    _peso(r['commission']),
                    _peso(r['net_to_seller']),
                    cat_para,
                ])

            # Totals row at the bottom
            total_subtotal = sum(r['subtotal'] for r in data['rows'])
            tx_rows.append([
                Paragraph("<b>TOTAL</b>", styles['small']),
                "",
                "",
                "",
                f"{summary['total_items']:,}",
                _peso(total_subtotal),
                _peso(summary['total_shipping']),
                _peso(summary['total_sales']),
                _peso(summary['total_commission']),
                _peso(summary['net_to_sellers']),
                "",
            ])

            tx_table = _pdf_make_table(
                headers=[
                    "Order No.", "Date / Time", "Seller", "Buyer", "Items",
                    "Subtotal", "Shipping", "Total", "Commission", "Net to Seller", "Categories",
                ],
                rows=tx_rows,
                col_widths=[
                    1.0 * inch, 0.95 * inch, 1.6 * inch, 1.4 * inch, 0.55 * inch,
                    0.95 * inch, 0.85 * inch, 0.95 * inch, 1.05 * inch, 1.10 * inch, 1.20 * inch,
                ],
                align_overrides={
                    0: 'CENTER', 1: 'CENTER', 4: 'CENTER',
                    5: 'RIGHT', 6: 'RIGHT', 7: 'RIGHT', 8: 'RIGHT', 9: 'RIGHT',
                },
            )

            # Make the totals row stand out
            tx_table.setStyle(TableStyle([
                ('BACKGROUND', (0, -1), (-1, -1), _PDF_BRAND),
                ('TEXTCOLOR', (0, -1), (-1, -1), _PDF_HEADER_TEXT),
                ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
            ]))
            flow.append(tx_table)

        flow.append(Spacer(1, 14))
        flow.append(Paragraph(
            "End of report • Generated automatically • Amounts in Philippine Peso",
            styles['footer'],
        ))

        # Build PDF with the page footer
        generated_at = data['generated_at']
        doc.build(
            flow,
            onFirstPage=lambda c, d: _on_page(c, d, generated_at),
            onLaterPages=lambda c, d: _on_page(c, d, generated_at),
        )

        buffer.seek(0)
        filename = f"velare_sales_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
        response = make_response(buffer.getvalue())
        response.headers['Content-Type'] = 'application/pdf'
        response.headers['Content-Disposition'] = f'attachment; filename={filename}'
        print(f"✅ PDF export completed ({summary['total_orders']} rows)")
        return response

    except Exception as e:
        print(f"❌ Error exporting PDF: {e}")
        import traceback
        traceback.print_exc()
        return f"Error exporting PDF: {str(e)}", 500
