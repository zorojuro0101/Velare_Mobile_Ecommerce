from flask import Blueprint, render_template, session, jsonify, request, make_response
import os
import sys
from datetime import datetime, timedelta
from decimal import Decimal
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from io import BytesIO

# Add the parent directory to the path to import db_config
sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
from database.db_config import get_supabase_client
from utils.auth_decorators import seller_required

seller_product_sales_bp = Blueprint('seller_product_sales', __name__)

@seller_product_sales_bp.route('/seller/product-sales')
@seller_required
def seller_product_sales():
    """Display seller product sales page"""
    try:
        seller_id = session.get('seller_id')
        print(f"🔍 Product Sales - seller_id: {seller_id}")
        
        supabase = get_supabase_client()
        if not supabase:
            print("❌ Supabase client not available")
            return render_template('seller/seller_product_sales.html', error='Database connection failed')
        
        # Get seller information for profile display
        seller_response = supabase.table('sellers').select(
            'first_name, last_name, shop_name, shop_logo'
        ).eq('seller_id', seller_id).execute()
        
        seller_info = seller_response.data[0] if seller_response.data else None
        print(f"👤 Seller info: {seller_info}")
        
        return render_template('seller/seller_product_sales.html', seller=seller_info)
        
    except Exception as e:
        print(f"❌ Error loading seller product sales: {e}")
        import traceback
        traceback.print_exc()
        return render_template('seller/seller_product_sales.html', error=str(e))

@seller_product_sales_bp.route('/seller/api/sales-summary')
@seller_required
def get_sales_summary():
    """API endpoint to get sales summary data"""
    try:
        seller_id = session.get('seller_id')
        date_from = request.args.get('dateFrom')
        date_to = request.args.get('dateTo')
        print(f"🔍 Sales Summary - seller_id: {seller_id}, date_from: {date_from}, date_to: {date_to}")
        
        supabase = get_supabase_client()
        if not supabase:
            print("❌ Supabase client not available")
            return jsonify({'error': 'Database connection failed'}), 500
        
        # Calculate date range
        if date_from and date_to:
            start_date = datetime.strptime(date_from, '%Y-%m-%d')
            end_date = datetime.strptime(date_to, '%Y-%m-%d')
            end_date = end_date.replace(hour=23, minute=59, second=59)
        else:
            end_date = datetime.now()
            start_date = end_date - timedelta(days=30)
        
        # Get orders for this seller in date range
        orders_query = supabase.table('orders').select(
            'order_id, commission_amount, order_items(quantity, subtotal, product_id)'
        ).eq('seller_id', seller_id).eq('order_status', 'delivered').eq('order_received', True).gte('created_at', start_date.isoformat()).lte('created_at', end_date.isoformat()).execute()
        
        orders = orders_query.data if orders_query.data else []
        print(f"📦 Found {len(orders)} delivered orders")
        
        # Calculate totals
        total_products_sold = 0
        total_revenue = 0
        total_commission = 0
        product_sales = {}  # Track sales per product
        
        for order in orders:
            total_commission += float(order.get('commission_amount', 0))
            order_items = order.get('order_items', [])
            
            for item in order_items:
                quantity = int(item.get('quantity', 0))
                subtotal = float(item.get('subtotal', 0))
                product_id = item.get('product_id')
                
                total_products_sold += quantity
                total_revenue += subtotal
                
                # Track per product for best seller
                if product_id:
                    if product_id not in product_sales:
                        product_sales[product_id] = {'units': 0, 'revenue': 0}
                    product_sales[product_id]['units'] += quantity
                    product_sales[product_id]['revenue'] += subtotal
        
        # Find best-selling product
        best_seller = None
        if product_sales:
            best_product_id = max(product_sales, key=lambda x: product_sales[x]['units'])
            product_response = supabase.table('products').select('product_name').eq('product_id', best_product_id).execute()
            
            if product_response.data:
                best_seller = {
                    'product_name': product_response.data[0]['product_name'],
                    'units_sold': product_sales[best_product_id]['units'],
                    'revenue': product_sales[best_product_id]['revenue']
                }
        
        print(f"📊 Summary: products={total_products_sold}, revenue={total_revenue:.2f}, commission={total_commission:.2f}")
        
        return jsonify({
            'total_products_sold': int(total_products_sold),
            'total_revenue': float(total_revenue),
            'total_commission': float(total_commission),
            'best_seller': best_seller
        })
        
    except Exception as e:
        print(f"❌ Error getting sales summary: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500

@seller_product_sales_bp.route('/seller/api/sales-details')
@seller_required
def get_sales_details():
    """API endpoint to get detailed sales data"""
    try:
        seller_id = session.get('seller_id')
        date_from = request.args.get('dateFrom')
        date_to = request.args.get('dateTo')
        sort_by = request.args.get('sort', 'most-sold')
        print(f"🔍 Sales Details - seller_id: {seller_id}, sort: {sort_by}")
        
        supabase = get_supabase_client()
        if not supabase:
            print("❌ Supabase client not available")
            return jsonify({'error': 'Database connection failed'}), 500
        
        # Calculate date range
        if date_from and date_to:
            start_date = datetime.strptime(date_from, '%Y-%m-%d')
            end_date = datetime.strptime(date_to, '%Y-%m-%d')
            end_date = end_date.replace(hour=23, minute=59, second=59)
        else:
            end_date = datetime.now()
            start_date = end_date - timedelta(days=30)
        
        # Get orders with items
        orders_query = supabase.table('orders').select(
            'order_id, created_at, order_items(product_id, quantity, subtotal, products(product_name))'
        ).eq('seller_id', seller_id).in_('order_status', ['in_transit', 'delivered']).gte('created_at', start_date.isoformat()).lte('created_at', end_date.isoformat()).execute()
        
        orders = orders_query.data if orders_query.data else []
        print(f"📦 Found {len(orders)} orders")
        
        # Aggregate by product
        product_sales = {}
        for order in orders:
            order_date = datetime.fromisoformat(order['created_at'].replace('Z', '+00:00'))
            order_items = order.get('order_items', [])
            
            for item in order_items:
                product_id = item.get('product_id')
                product_data = item.get('products', {})
                product_name = product_data.get('product_name', 'Unknown') if product_data else 'Unknown'
                quantity = int(item.get('quantity', 0))
                subtotal = float(item.get('subtotal', 0))
                
                if product_id not in product_sales:
                    product_sales[product_id] = {
                        'product_name': product_name,
                        'units_sold': 0,
                        'revenue': 0,
                        'first_order': order_date,
                        'latest_order': order_date
                    }
                
                product_sales[product_id]['units_sold'] += quantity
                product_sales[product_id]['revenue'] += subtotal
                
                if order_date < product_sales[product_id]['first_order']:
                    product_sales[product_id]['first_order'] = order_date
                if order_date > product_sales[product_id]['latest_order']:
                    product_sales[product_id]['latest_order'] = order_date
        
        # Convert to list and sort
        sales_list = list(product_sales.values())
        
        if sort_by == 'highest-revenue':
            sales_list.sort(key=lambda x: x['revenue'], reverse=True)
        elif sort_by == 'latest':
            sales_list.sort(key=lambda x: x['latest_order'], reverse=True)
        else:  # most-sold
            sales_list.sort(key=lambda x: x['units_sold'], reverse=True)
        
        # Format the data
        formatted_sales = []
        for sale in sales_list:
            formatted_sales.append({
                'product_name': sale['product_name'],
                'units_sold': int(sale['units_sold']),
                'revenue': float(sale['revenue']),
                'date_range': f"{sale['first_order'].strftime('%b %d')} - {sale['latest_order'].strftime('%b %d, %Y')}"
            })
        
        print(f"📊 Returning {len(formatted_sales)} products")
        return jsonify({'sales': formatted_sales})
        
    except Exception as e:
        print(f"❌ Error getting sales details: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500

@seller_product_sales_bp.route('/seller/api/sales-charts')
@seller_required
def get_sales_charts():
    """API endpoint to get chart data for sales analytics based on date range"""
    try:
        seller_id = session.get('seller_id')
        date_from = request.args.get('dateFrom')
        date_to = request.args.get('dateTo')
        print(f"🔍 Sales Charts - seller_id: {seller_id}")
        
        supabase = get_supabase_client()
        if not supabase:
            print("❌ Supabase client not available")
            return jsonify({'error': 'Database connection failed'}), 500
        
        # Calculate date range
        if date_from and date_to:
            start_date = datetime.strptime(date_from, '%Y-%m-%d')
            end_date = datetime.strptime(date_to, '%Y-%m-%d')
            end_date = end_date.replace(hour=23, minute=59, second=59)
        else:
            end_date = datetime.now()
            start_date = end_date - timedelta(days=30)
        
        days_diff = (end_date - start_date).days
        print(f"📅 Date range: {days_diff} days")
        
        # Get all orders with items in date range
        orders_query = supabase.table('orders').select(
            'order_id, created_at, order_items(quantity, subtotal)'
        ).eq('seller_id', seller_id).in_('order_status', ['in_transit', 'delivered']).gte('created_at', start_date.isoformat()).lte('created_at', end_date.isoformat()).execute()
        
        orders = orders_query.data if orders_query.data else []
        print(f"📦 Found {len(orders)} orders for charts")
        
        # Process data based on grouping
        units_data = []
        revenue_data = []
        units_label = ''
        revenue_label = ''
        
        if days_diff <= 7:
            # Daily grouping
            units_label = 'Daily Units Sold'
            revenue_label = 'Daily Revenue'
            daily_units = {}
            daily_revenue = {}
            
            for order in orders:
                order_date = datetime.fromisoformat(order['created_at'].replace('Z', '+00:00'))
                date_key = order_date.strftime('%b %d')
                
                if date_key not in daily_units:
                    daily_units[date_key] = 0
                    daily_revenue[date_key] = 0
                
                order_items = order.get('order_items', [])
                for item in order_items:
                    daily_units[date_key] += int(item.get('quantity', 0))
                    daily_revenue[date_key] += float(item.get('subtotal', 0))
            
            # Sort by date
            for date_label in sorted(daily_units.keys(), key=lambda x: datetime.strptime(x, '%b %d')):
                units_data.append({'date_label': date_label, 'units_sold': daily_units[date_label]})
                revenue_data.append({'date_label': date_label, 'revenue': daily_revenue[date_label]})
                
        elif days_diff <= 30:
            # Weekly grouping
            units_label = 'Weekly Units Sold'
            revenue_label = 'Weekly Revenue'
            weekly_units = {}
            weekly_revenue = {}
            
            for order in orders:
                order_date = datetime.fromisoformat(order['created_at'].replace('Z', '+00:00'))
                week_start = order_date - timedelta(days=order_date.weekday())
                date_key = week_start.strftime('%b %d')
                
                if date_key not in weekly_units:
                    weekly_units[date_key] = 0
                    weekly_revenue[date_key] = 0
                
                order_items = order.get('order_items', [])
                for item in order_items:
                    weekly_units[date_key] += int(item.get('quantity', 0))
                    weekly_revenue[date_key] += float(item.get('subtotal', 0))
            
            # Sort by date
            for date_label in sorted(weekly_units.keys(), key=lambda x: datetime.strptime(x, '%b %d')):
                units_data.append({'date_label': date_label, 'units_sold': weekly_units[date_label]})
                revenue_data.append({'date_label': date_label, 'revenue': weekly_revenue[date_label]})
                
        else:
            # Monthly grouping
            units_label = 'Monthly Units Sold'
            revenue_label = 'Monthly Revenue'
            monthly_units = {}
            monthly_revenue = {}
            
            for order in orders:
                order_date = datetime.fromisoformat(order['created_at'].replace('Z', '+00:00'))
                date_key = order_date.strftime('%b %Y')
                
                if date_key not in monthly_units:
                    monthly_units[date_key] = 0
                    monthly_revenue[date_key] = 0
                
                order_items = order.get('order_items', [])
                for item in order_items:
                    monthly_units[date_key] += int(item.get('quantity', 0))
                    monthly_revenue[date_key] += float(item.get('subtotal', 0))
            
            # Sort by date
            for date_label in sorted(monthly_units.keys(), key=lambda x: datetime.strptime(x, '%b %Y')):
                units_data.append({'date_label': date_label, 'units_sold': monthly_units[date_label]})
                revenue_data.append({'date_label': date_label, 'revenue': monthly_revenue[date_label]})
        
        print(f"📊 Chart data: {len(units_data)} data points")
        
        return jsonify({
            'units': {
                'labels': [u['date_label'] for u in units_data] if units_data else [],
                'data': [int(u['units_sold']) for u in units_data] if units_data else [],
                'label': units_label
            },
            'revenue': {
                'labels': [r['date_label'] for r in revenue_data] if revenue_data else [],
                'data': [float(r['revenue']) for r in revenue_data] if revenue_data else [],
                'label': revenue_label
            }
        })
        
    except Exception as e:
        print(f"❌ Error getting chart data: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500


@seller_product_sales_bp.route('/seller/product-sales/export-excel')
@seller_required
def export_excel():
    """Export seller product sales to Excel format"""
    try:
        seller_id = session.get('seller_id')
        date_from = request.args.get('dateFrom')
        date_to = request.args.get('dateTo')
        
        print(f"📊 Excel Export Request - Seller ID: {seller_id}, Date From: {date_from}, Date To: {date_to}")
        
        supabase = get_supabase_client()
        if not supabase:
            print("❌ Supabase client not available")
            return "Database connection failed", 500
        
        # Get seller info
        seller_response = supabase.table('sellers').select('shop_name, first_name, last_name').eq('seller_id', seller_id).execute()
        seller_info = seller_response.data[0] if seller_response.data else {'shop_name': 'Shop', 'first_name': '', 'last_name': ''}
        
        # Calculate date range
        if date_from and date_to:
            start_date = datetime.strptime(date_from, '%Y-%m-%d')
            end_date = datetime.strptime(date_to, '%Y-%m-%d')
            end_date = end_date.replace(hour=23, minute=59, second=59)
        else:
            end_date = datetime.now()
            start_date = end_date - timedelta(days=30)
        
        # Get orders for summary
        orders_query = supabase.table('orders').select(
            'order_id, commission_amount, order_items(quantity, subtotal, product_id, product_name, products(product_name))'
        ).eq('seller_id', seller_id).eq('order_status', 'delivered').eq('order_received', True).gte('created_at', start_date.isoformat()).lte('created_at', end_date.isoformat()).execute()
        
        orders = orders_query.data if orders_query.data else []
        
        # Calculate summary
        total_orders = len(orders)
        total_products_sold = 0
        total_revenue = 0
        total_commission = 0
        product_details = {}
        
        for order in orders:
            total_commission += float(order.get('commission_amount', 0))
            order_items = order.get('order_items', [])
            
            for item in order_items:
                quantity = int(item.get('quantity', 0))
                subtotal = float(item.get('subtotal', 0))
                product_id = item.get('product_id')
                product_name = item.get('product_name', 'Unknown')
                
                total_products_sold += quantity
                total_revenue += subtotal
                
                if product_id not in product_details:
                    product_details[product_id] = {
                        'product_name': product_name,
                        'units_sold': 0,
                        'revenue': 0,
                        'order_count': set()
                    }
                
                product_details[product_id]['units_sold'] += quantity
                product_details[product_id]['revenue'] += subtotal
                product_details[product_id]['order_count'].add(order['order_id'])
        
        # Get detailed sales with dates
        orders_with_dates = supabase.table('orders').select(
            'order_id, created_at, order_items(product_id, product_name, quantity, subtotal)'
        ).eq('seller_id', seller_id).eq('order_status', 'delivered').eq('order_received', True).gte('created_at', start_date.isoformat()).lte('created_at', end_date.isoformat()).execute()
        
        product_dates = {}
        for order in orders_with_dates.data if orders_with_dates.data else []:
            order_date = datetime.fromisoformat(order['created_at'].replace('Z', '+00:00'))
            for item in order.get('order_items', []):
                product_id = item.get('product_id')
                if product_id:
                    if product_id not in product_dates:
                        product_dates[product_id] = {'first': order_date, 'latest': order_date}
                    else:
                        if order_date < product_dates[product_id]['first']:
                            product_dates[product_id]['first'] = order_date
                        if order_date > product_dates[product_id]['latest']:
                            product_dates[product_id]['latest'] = order_date
        
        # Format sales details
        sales_details = []
        for product_id, details in product_details.items():
            sales_details.append({
                'product_name': details['product_name'],
                'units_sold': details['units_sold'],
                'revenue': details['revenue'],
                'order_count': len(details['order_count']),
                'first_order': product_dates.get(product_id, {}).get('first', datetime.now()),
                'latest_order': product_dates.get(product_id, {}).get('latest', datetime.now())
            })
        
        sales_details.sort(key=lambda x: x['units_sold'], reverse=True)
        
        # Get time-based breakdown
        days_diff = (end_date - start_date).days
        time_breakdown = []
        breakdown_type = "Weekly" if days_diff <= 30 else "Monthly"
        
        if days_diff <= 30:
            # Weekly breakdown
            weekly_data = {}
            for order in orders_with_dates.data if orders_with_dates.data else []:
                order_date = datetime.fromisoformat(order['created_at'].replace('Z', '+00:00'))
                week_start = order_date - timedelta(days=order_date.weekday())
                date_key = week_start.strftime('%b %d')
                
                if date_key not in weekly_data:
                    weekly_data[date_key] = {'units': 0, 'revenue': 0, 'orders': set()}
                
                weekly_data[date_key]['orders'].add(order['order_id'])
                for item in order.get('order_items', []):
                    weekly_data[date_key]['units'] += int(item.get('quantity', 0))
                    weekly_data[date_key]['revenue'] += float(item.get('subtotal', 0))
            
            for date_label in sorted(weekly_data.keys(), key=lambda x: datetime.strptime(x, '%b %d')):
                time_breakdown.append({
                    'period_label': date_label,
                    'order_count': len(weekly_data[date_label]['orders']),
                    'units_sold': weekly_data[date_label]['units'],
                    'revenue': weekly_data[date_label]['revenue']
                })
        else:
            # Monthly breakdown
            monthly_data = {}
            for order in orders_with_dates.data if orders_with_dates.data else []:
                order_date = datetime.fromisoformat(order['created_at'].replace('Z', '+00:00'))
                date_key = order_date.strftime('%b %Y')
                
                if date_key not in monthly_data:
                    monthly_data[date_key] = {'units': 0, 'revenue': 0, 'orders': set()}
                
                monthly_data[date_key]['orders'].add(order['order_id'])
                for item in order.get('order_items', []):
                    monthly_data[date_key]['units'] += int(item.get('quantity', 0))
                    monthly_data[date_key]['revenue'] += float(item.get('subtotal', 0))
            
            for date_label in sorted(monthly_data.keys(), key=lambda x: datetime.strptime(x, '%b %Y')):
                time_breakdown.append({
                    'period_label': date_label,
                    'order_count': len(monthly_data[date_label]['orders']),
                    'units_sold': monthly_data[date_label]['units'],
                    'revenue': monthly_data[date_label]['revenue']
                })
        
        # Get Product Sold data
        sold_orders_query = supabase.table('orders').select(
            'order_id, order_number, subtotal, commission_amount, updated_at, buyer_id, buyers(first_name, last_name), order_items(product_name, variant_color, variant_size, quantity, unit_price, subtotal)'
        ).eq('seller_id', seller_id).eq('order_status', 'delivered').eq('order_received', True).gte('created_at', start_date.isoformat()).lte('created_at', end_date.isoformat()).order('updated_at', desc=True).execute()
        
        sold_orders = []
        for order in sold_orders_query.data if sold_orders_query.data else []:
            buyer_data = order.get('buyers', {})
            sold_orders.append({
                'order_id': order['order_id'],
                'order_number': order.get('order_number'),
                'order_total': float(order.get('subtotal', 0)),
                'commission_amount': float(order.get('commission_amount', 0)),
                'order_received_date': datetime.fromisoformat(order['updated_at'].replace('Z', '+00:00')),
                'buyer_name': f"{buyer_data.get('first_name', '')} {buyer_data.get('last_name', '')}".strip() if buyer_data else 'Unknown',
                'items': order.get('order_items', [])
            })
        
        # Create Excel workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Product Sales"
        
        # Define styles
        header_font = Font(name='Calibri', size=18, bold=True, color="1F4E78")
        subheader_font = Font(name='Calibri', size=14, bold=True, color="1F4E78")
        summary_header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        summary_header_font = Font(name='Calibri', size=11, bold=True, color="FFFFFF")
        table_header_fill = PatternFill(start_color="5B9BD5", end_color="5B9BD5", fill_type="solid")
        table_header_font = Font(name='Calibri', size=11, bold=True, color="FFFFFF")
        alt_row_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        
        border = Border(
            left=Side(style='thin', color='000000'),
            right=Side(style='thin', color='000000'),
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000')
        )
        
        center_align = Alignment(horizontal='center', vertical='center')
        left_align = Alignment(horizontal='left', vertical='center')
        right_align = Alignment(horizontal='right', vertical='center')
        
        current_row = 1
        
        # Header
        ws.merge_cells(f'A{current_row}:F{current_row}')
        cell = ws[f'A{current_row}']
        cell.value = seller_info['shop_name']
        cell.font = header_font
        cell.alignment = center_align
        current_row += 1
        
        ws.merge_cells(f'A{current_row}:F{current_row}')
        cell = ws[f'A{current_row}']
        cell.value = "PRODUCT SALES REPORT"
        cell.font = subheader_font
        cell.alignment = center_align
        current_row += 2
        
        # Report info
        ws[f'A{current_row}'] = "Report Generated:"
        ws[f'A{current_row}'].font = Font(bold=True)
        ws[f'B{current_row}'] = datetime.now().strftime('%B %d, %Y at %I:%M %p')
        current_row += 1
        
        ws[f'A{current_row}'] = "Report Period:"
        ws[f'A{current_row}'].font = Font(bold=True)
        ws[f'B{current_row}'] = f"{start_date.strftime('%B %d, %Y')} to {end_date.strftime('%B %d, %Y')}"
        current_row += 2
        
        # Summary section
        ws.merge_cells(f'A{current_row}:B{current_row}')
        cell = ws[f'A{current_row}']
        cell.value = "SALES SUMMARY"
        cell.fill = summary_header_fill
        cell.font = summary_header_font
        cell.alignment = center_align
        cell.border = border
        ws[f'B{current_row}'].border = border
        current_row += 1
        
        # Summary data
        summary_data = [
            ("Total Orders", total_orders),
            ("Total Products Sold", total_products_sold),
            ("Total Revenue", f"₱{total_revenue:,.2f}"),
            ("Platform Commission", f"₱{total_commission:,.2f}"),
            ("Net Earnings", f"₱{(total_revenue - total_commission):,.2f}")
        ]
        
        for label, value in summary_data:
            ws[f'A{current_row}'] = label
            ws[f'A{current_row}'].font = Font(bold=True)
            ws[f'A{current_row}'].border = border
            ws[f'B{current_row}'] = value
            ws[f'B{current_row}'].border = border
            ws[f'B{current_row}'].alignment = right_align
            current_row += 1
        
        current_row += 1
        
        # Product sales table
        ws.merge_cells(f'A{current_row}:F{current_row}')
        cell = ws[f'A{current_row}']
        cell.value = "PRODUCT SALES DETAILS"
        cell.fill = summary_header_fill
        cell.font = summary_header_font
        cell.alignment = center_align
        for col in ['A', 'B', 'C', 'D', 'E', 'F']:
            ws[f'{col}{current_row}'].border = border
        current_row += 1
        
        # Table headers
        headers = ['Product Name', 'Units Sold', 'Revenue', 'Orders', 'First Sale', 'Latest Sale']
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=current_row, column=col_idx)
            cell.value = header
            cell.fill = table_header_fill
            cell.font = table_header_font
            cell.alignment = center_align
            cell.border = border
        current_row += 1
        
        # Table data
        for idx, sale in enumerate(sales_details):
            row_fill = alt_row_fill if idx % 2 == 0 else None
            
            cell = ws.cell(row=current_row, column=1, value=sale['product_name'])
            cell.border = border
            if row_fill:
                cell.fill = row_fill
            cell.alignment = left_align
            
            cell = ws.cell(row=current_row, column=2, value=int(sale['units_sold']))
            cell.border = border
            if row_fill:
                cell.fill = row_fill
            cell.alignment = center_align
            
            cell = ws.cell(row=current_row, column=3, value=f"₱{float(sale['revenue']):,.2f}")
            cell.border = border
            if row_fill:
                cell.fill = row_fill
            cell.alignment = right_align
            
            cell = ws.cell(row=current_row, column=4, value=int(sale['order_count']))
            cell.border = border
            if row_fill:
                cell.fill = row_fill
            cell.alignment = center_align
            
            cell = ws.cell(row=current_row, column=5, value=sale['first_order'].strftime('%b %d, %Y'))
            cell.border = border
            if row_fill:
                cell.fill = row_fill
            cell.alignment = center_align
            
            cell = ws.cell(row=current_row, column=6, value=sale['latest_order'].strftime('%b %d, %Y'))
            cell.border = border
            if row_fill:
                cell.fill = row_fill
            cell.alignment = center_align
            
            current_row += 1
        
        current_row += 2
        
        # TIME-BASED BREAKDOWN TABLE
        ws.merge_cells(f'A{current_row}:E{current_row}')
        cell = ws[f'A{current_row}']
        cell.value = f"{breakdown_type.upper()} SALES BREAKDOWN"
        cell.fill = summary_header_fill
        cell.font = summary_header_font
        cell.alignment = center_align
        for col in ['A', 'B', 'C', 'D', 'E']:
            ws[f'{col}{current_row}'].border = border
        current_row += 1
        
        # Time breakdown table headers
        time_headers = ['Period', 'Orders', 'Units Sold', 'Revenue', 'Avg Order Value']
        for col_idx, header in enumerate(time_headers, start=1):
            cell = ws.cell(row=current_row, column=col_idx)
            cell.value = header
            cell.fill = table_header_fill
            cell.font = table_header_font
            cell.alignment = center_align
            cell.border = border
        current_row += 1
        
        # Time breakdown data
        for idx, period in enumerate(time_breakdown):
            row_fill = alt_row_fill if idx % 2 == 0 else None
            avg_order_value = float(period['revenue']) / int(period['order_count']) if period['order_count'] > 0 else 0
            
            cell = ws.cell(row=current_row, column=1, value=period['period_label'])
            cell.border = border
            if row_fill:
                cell.fill = row_fill
            cell.alignment = left_align
            
            cell = ws.cell(row=current_row, column=2, value=int(period['order_count']))
            cell.border = border
            if row_fill:
                cell.fill = row_fill
            cell.alignment = center_align
            
            cell = ws.cell(row=current_row, column=3, value=int(period['units_sold']))
            cell.border = border
            if row_fill:
                cell.fill = row_fill
            cell.alignment = center_align
            
            cell = ws.cell(row=current_row, column=4, value=f"₱{float(period['revenue']):,.2f}")
            cell.border = border
            if row_fill:
                cell.fill = row_fill
            cell.alignment = right_align
            
            cell = ws.cell(row=current_row, column=5, value=f"₱{avg_order_value:,.2f}")
            cell.border = border
            if row_fill:
                cell.fill = row_fill
            cell.alignment = right_align
            
            current_row += 1
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 40
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 18
        ws.column_dimensions['D'].width = 12
        ws.column_dimensions['E'].width = 18
        ws.column_dimensions['F'].width = 18
        
        # PRODUCT SOLD SHEET
        ws2 = wb.create_sheet(title="Product Sold")
        
        current_row = 1
        
        # Header
        ws2.merge_cells(f'A{current_row}:F{current_row}')
        cell = ws2[f'A{current_row}']
        cell.value = seller_info['shop_name']
        cell.font = header_font
        cell.alignment = center_align
        current_row += 1
        
        ws2.merge_cells(f'A{current_row}:F{current_row}')
        cell = ws2[f'A{current_row}']
        cell.value = "PRODUCT SOLD REPORT"
        cell.font = subheader_font
        cell.alignment = center_align
        current_row += 2
        
        # Report info
        ws2[f'A{current_row}'] = "Report Generated:"
        ws2[f'A{current_row}'].font = Font(bold=True)
        ws2[f'B{current_row}'] = datetime.now().strftime('%B %d, %Y at %I:%M %p')
        current_row += 1
        
        ws2[f'A{current_row}'] = "Report Period:"
        ws2[f'A{current_row}'].font = Font(bold=True)
        ws2[f'B{current_row}'] = f"{start_date.strftime('%B %d, %Y')} to {end_date.strftime('%B %d, %Y')}"
        current_row += 2
        
        # Product Sold table
        ws2.merge_cells(f'A{current_row}:F{current_row}')
        cell = ws2[f'A{current_row}']
        cell.value = "PRODUCT SOLD DETAILS"
        cell.fill = summary_header_fill
        cell.font = summary_header_font
        cell.alignment = center_align
        for col in ['A', 'B', 'C', 'D', 'E', 'F']:
            ws2[f'{col}{current_row}'].border = border
        current_row += 1
        
        # Table headers
        sold_headers = ['#', 'Products', 'Order Received', 'Buyer', 'Order Total', 'Commission (5%)']
        for col_idx, header in enumerate(sold_headers, start=1):
            cell = ws2.cell(row=current_row, column=col_idx)
            cell.value = header
            cell.fill = table_header_fill
            cell.font = table_header_font
            cell.alignment = center_align
            cell.border = border
        current_row += 1
        
        # Table data
        for idx, order in enumerate(sold_orders, start=1):
            row_fill = alt_row_fill if idx % 2 == 0 else None
            
            # Order number
            cell = ws2.cell(row=current_row, column=1, value=f"#{idx:03d}")
            cell.border = border
            if row_fill:
                cell.fill = row_fill
            cell.alignment = center_align
            
            # Products (with variants and quantities)
            products_text = []
            for item in order['items']:
                variant_parts = []
                if item.get('variant_color'):
                    variant_parts.append(item['variant_color'])
                if item.get('variant_size'):
                    variant_parts.append(item['variant_size'])
                variant_text = f" ({', '.join(variant_parts)})" if variant_parts else ""
                products_text.append(f"{item.get('product_name', 'Unknown')}{variant_text} × {item.get('quantity', 0)}")
            
            cell = ws2.cell(row=current_row, column=2, value="\n".join(products_text))
            cell.border = border
            if row_fill:
                cell.fill = row_fill
            cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
            
            # Order Received Date
            cell = ws2.cell(row=current_row, column=3, value=order['order_received_date'].strftime('%b %d, %Y'))
            cell.border = border
            if row_fill:
                cell.fill = row_fill
            cell.alignment = center_align
            
            # Buyer
            cell = ws2.cell(row=current_row, column=4, value=order['buyer_name'])
            cell.border = border
            if row_fill:
                cell.fill = row_fill
            cell.alignment = left_align
            
            # Order Total
            cell = ws2.cell(row=current_row, column=5, value=f"₱{float(order['order_total']):,.2f}")
            cell.border = border
            if row_fill:
                cell.fill = row_fill
            cell.alignment = right_align
            cell.font = Font(color="008000", bold=True)  # Green for positive
            
            # Commission
            cell = ws2.cell(row=current_row, column=6, value=f"₱{float(order['commission_amount']):,.2f}")
            cell.border = border
            if row_fill:
                cell.fill = row_fill
            cell.alignment = right_align
            cell.font = Font(color="FF0000", bold=True)  # Red for negative
            
            # Set row height for wrapped text
            ws2.row_dimensions[current_row].height = 15 * len(order['items'])
            
            current_row += 1
        
        # Adjust column widths for Product Sold sheet
        ws2.column_dimensions['A'].width = 8
        ws2.column_dimensions['B'].width = 50
        ws2.column_dimensions['C'].width = 18
        ws2.column_dimensions['D'].width = 25
        ws2.column_dimensions['E'].width = 18
        ws2.column_dimensions['F'].width = 18
        
        # Save to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Generate filename
        shop_name_safe = seller_info['shop_name'].replace(' ', '_').replace("'", "")
        filename = f"Product_Sales_{shop_name_safe}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        print(f"✅ Excel file created successfully: {filename}")
        
        # Create response
        response = make_response(output.getvalue())
        response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        response.headers['Content-Disposition'] = f'attachment; filename={filename}'
        
        return response
        
    except Exception as e:
        print(f"❌ Error exporting to Excel: {e}")
        import traceback
        traceback.print_exc()
        return f"Error exporting to Excel: {str(e)}", 500

