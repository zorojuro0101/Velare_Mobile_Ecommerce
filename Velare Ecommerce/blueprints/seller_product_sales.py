from flask import Blueprint, render_template, session, jsonify, request, make_response
import os
import sys
from datetime import datetime, timedelta
from decimal import Decimal
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from io import BytesIO

# Add the parent directory to the path to import db_config
sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
from database.db_config import get_db_connection, close_db_connection
from utils.auth_decorators import seller_required

seller_product_sales_bp = Blueprint('seller_product_sales', __name__)

@seller_product_sales_bp.route('/seller/product-sales')
@seller_required
def seller_product_sales():
    """Display seller product sales page"""
    try:
        seller_id = session.get('seller_id')
        
        connection = get_db_connection()
        if not connection:
            return render_template('seller/seller_product_sales.html', error='Database connection failed')
        
        cursor = connection.cursor(dictionary=True)
        
        # Get seller information for profile display
        cursor.execute("""
            SELECT first_name, last_name, shop_name, shop_logo
            FROM sellers
            WHERE seller_id = %s
        """, (seller_id,))
        seller_info = cursor.fetchone()
        
        # Fix shop_logo path: remove 'static/' prefix for url_for
        if seller_info and seller_info.get('shop_logo'):
            if seller_info['shop_logo'].startswith('static/'):
                seller_info['shop_logo'] = seller_info['shop_logo'][7:]  # Remove 'static/' prefix
        
        close_db_connection(connection, cursor)
        
        return render_template('seller/seller_product_sales.html', seller=seller_info)
        
    except Exception as e:
        print(f"Error loading seller product sales: {e}")
        return render_template('seller/seller_product_sales.html', error=str(e))

@seller_product_sales_bp.route('/seller/api/sales-summary')
@seller_required
def get_sales_summary():
    """API endpoint to get sales summary data"""
    try:
        seller_id = session.get('seller_id')
        date_from = request.args.get('dateFrom')
        date_to = request.args.get('dateTo')
        
        connection = get_db_connection()
        if not connection:
            return jsonify({'error': 'Database connection failed'}), 500
        
        cursor = connection.cursor(dictionary=True)
        
        # Calculate date range
        if date_from and date_to:
            # Custom date range provided
            start_date = datetime.strptime(date_from, '%Y-%m-%d')
            end_date = datetime.strptime(date_to, '%Y-%m-%d')
            # Set end_date to end of day
            end_date = end_date.replace(hour=23, minute=59, second=59)
        else:
            # Default to last 30 days
            end_date = datetime.now()
            start_date = end_date - timedelta(days=30)
        
        # Get total products sold and revenue from order_items
        cursor.execute("""
            SELECT 
                COALESCE(SUM(oi.quantity), 0) as total_products_sold,
                COALESCE(SUM(oi.subtotal), 0) as total_revenue
            FROM orders o
            JOIN order_items oi ON o.order_id = oi.order_id
            WHERE o.seller_id = %s
                AND o.order_status = 'delivered'
                AND o.order_received = TRUE
                AND o.created_at >= %s
                AND o.created_at <= %s
        """, (seller_id, start_date, end_date))
        summary = cursor.fetchone()
        
        # Get total commission separately (without JOIN to avoid double counting)
        cursor.execute("""
            SELECT 
                COALESCE(SUM(commission_amount), 0) as total_commission
            FROM orders
            WHERE seller_id = %s
                AND order_status = 'delivered'
                AND order_received = TRUE
                AND created_at >= %s
                AND created_at <= %s
        """, (seller_id, start_date, end_date))
        commission_data = cursor.fetchone()
        
        # Get best-selling product (only from delivered and confirmed orders)
        cursor.execute("""
            SELECT 
                p.product_name,
                SUM(oi.quantity) as units_sold,
                SUM(oi.subtotal) as revenue
            FROM order_items oi
            JOIN orders o ON oi.order_id = o.order_id
            JOIN products p ON oi.product_id = p.product_id
            WHERE o.seller_id = %s
                AND o.order_status = 'delivered'
                AND o.order_received = TRUE
                AND o.created_at >= %s
                AND o.created_at <= %s
            GROUP BY oi.product_id, p.product_name
            ORDER BY units_sold DESC
            LIMIT 1
        """, (seller_id, start_date, end_date))
        best_seller = cursor.fetchone()
        
        close_db_connection(connection, cursor)
        
        return jsonify({
            'total_products_sold': int(summary['total_products_sold']) if summary else 0,
            'total_revenue': float(summary['total_revenue']) if summary else 0,
            'total_commission': float(commission_data['total_commission']) if commission_data else 0,
            'best_seller': {
                'product_name': best_seller['product_name'] if best_seller else 'N/A',
                'units_sold': int(best_seller['units_sold']) if best_seller else 0,
                'revenue': float(best_seller['revenue']) if best_seller else 0
            } if best_seller else None
        })
        
    except Exception as e:
        print(f"Error getting sales summary: {e}")
        return jsonify({'error': str(e)}), 500

@seller_product_sales_bp.route('/seller/api/sales-details')
@seller_required
def get_sales_details():
    """API endpoint to get detailed sales data"""
    try:
        seller_id = session.get('seller_id')
        date_from = request.args.get('dateFrom')
        date_to = request.args.get('dateTo')
        sort_by = request.args.get('sort', 'most-sold')
        
        connection = get_db_connection()
        if not connection:
            return jsonify({'error': 'Database connection failed'}), 500
        
        cursor = connection.cursor(dictionary=True)
        
        # Calculate date range
        if date_from and date_to:
            # Custom date range provided
            start_date = datetime.strptime(date_from, '%Y-%m-%d')
            end_date = datetime.strptime(date_to, '%Y-%m-%d')
            # Set end_date to end of day
            end_date = end_date.replace(hour=23, minute=59, second=59)
        else:
            # Default to last 30 days
            end_date = datetime.now()
            start_date = end_date - timedelta(days=30)
        
        # Determine sort order
        if sort_by == 'highest-revenue':
            order_clause = 'ORDER BY revenue DESC'
        elif sort_by == 'latest':
            order_clause = 'ORDER BY latest_order DESC'
        else:  # most-sold
            order_clause = 'ORDER BY units_sold DESC'
        
        # Get sales details per product
        query = f"""
            SELECT 
                p.product_name,
                SUM(oi.quantity) as units_sold,
                SUM(oi.subtotal) as revenue,
                MIN(o.created_at) as first_order,
                MAX(o.created_at) as latest_order
            FROM order_items oi
            JOIN orders o ON oi.order_id = o.order_id
            JOIN products p ON oi.product_id = p.product_id
            WHERE o.seller_id = %s
                AND o.order_status IN ('in_transit', 'delivered')
                AND o.created_at >= %s
                AND o.created_at <= %s
            GROUP BY oi.product_id, p.product_name
            {order_clause}
        """
        
        cursor.execute(query, (seller_id, start_date, end_date))
        sales_details = cursor.fetchall()
        
        # Format the data
        formatted_sales = []
        for sale in sales_details:
            formatted_sales.append({
                'product_name': sale['product_name'],
                'units_sold': int(sale['units_sold']),
                'revenue': float(sale['revenue']),
                'date_range': f"{sale['first_order'].strftime('%b %d')} - {sale['latest_order'].strftime('%b %d, %Y')}"
            })
        
        close_db_connection(connection, cursor)
        
        return jsonify({'sales': formatted_sales})
        
    except Exception as e:
        print(f"Error getting sales details: {e}")
        return jsonify({'error': str(e)}), 500

@seller_product_sales_bp.route('/seller/api/sales-charts')
@seller_required
def get_sales_charts():
    """API endpoint to get chart data for sales analytics based on date range"""
    try:
        seller_id = session.get('seller_id')
        date_from = request.args.get('dateFrom')
        date_to = request.args.get('dateTo')
        
        connection = get_db_connection()
        if not connection:
            return jsonify({'error': 'Database connection failed'}), 500
        
        cursor = connection.cursor(dictionary=True)
        
        # Calculate date range
        if date_from and date_to:
            start_date = datetime.strptime(date_from, '%Y-%m-%d')
            end_date = datetime.strptime(date_to, '%Y-%m-%d')
            end_date = end_date.replace(hour=23, minute=59, second=59)
        else:
            end_date = datetime.now()
            start_date = end_date - timedelta(days=30)
        
        # Calculate the number of days in the range
        days_diff = (end_date - start_date).days
        
        # Determine grouping based on date range
        if days_diff <= 7:
            # Daily grouping for 7 days or less
            cursor.execute("""
                SELECT 
                    DATE_FORMAT(o.created_at, '%b %d') as date_label,
                    SUM(oi.quantity) as units_sold
                FROM orders o
                JOIN order_items oi ON o.order_id = oi.order_id
                WHERE o.seller_id = %s
                    AND o.order_status IN ('in_transit', 'delivered')
                    AND o.created_at >= %s
                    AND o.created_at <= %s
                GROUP BY DATE(o.created_at), date_label
                ORDER BY DATE(o.created_at)
            """, (seller_id, start_date, end_date))
            units_data = cursor.fetchall()
            units_label = 'Daily Units Sold'
        elif days_diff <= 30:
            # Weekly grouping for up to 30 days
            cursor.execute("""
                SELECT 
                    DATE_FORMAT(o.created_at, '%b %d') as date_label,
                    SUM(oi.quantity) as units_sold
                FROM orders o
                JOIN order_items oi ON o.order_id = oi.order_id
                WHERE o.seller_id = %s
                    AND o.order_status IN ('in_transit', 'delivered')
                    AND o.created_at >= %s
                    AND o.created_at <= %s
                GROUP BY WEEK(o.created_at, 1), date_label
                ORDER BY WEEK(o.created_at, 1)
            """, (seller_id, start_date, end_date))
            units_data = cursor.fetchall()
            units_label = 'Weekly Units Sold'
        else:
            # Monthly grouping for longer periods (>30 days)
            cursor.execute("""
                SELECT 
                    DATE_FORMAT(o.created_at, '%b %Y') as date_label,
                    SUM(oi.quantity) as units_sold
                FROM orders o
                JOIN order_items oi ON o.order_id = oi.order_id
                WHERE o.seller_id = %s
                    AND o.order_status IN ('in_transit', 'delivered')
                    AND o.created_at >= %s
                    AND o.created_at <= %s
                GROUP BY YEAR(o.created_at), MONTH(o.created_at), date_label
                ORDER BY YEAR(o.created_at), MONTH(o.created_at)
            """, (seller_id, start_date, end_date))
            units_data = cursor.fetchall()
            units_label = 'Monthly Units Sold'
        
        # Get revenue trend with same grouping logic
        if days_diff <= 7:
            cursor.execute("""
                SELECT 
                    DATE_FORMAT(o.created_at, '%b %d') as date_label,
                    SUM(oi.subtotal) as revenue
                FROM orders o
                JOIN order_items oi ON o.order_id = oi.order_id
                WHERE o.seller_id = %s
                    AND o.order_status IN ('in_transit', 'delivered')
                    AND o.created_at >= %s
                    AND o.created_at <= %s
                GROUP BY DATE(o.created_at), date_label
                ORDER BY DATE(o.created_at)
            """, (seller_id, start_date, end_date))
            revenue_data = cursor.fetchall()
            revenue_label = 'Daily Revenue'
        elif days_diff <= 30:
            cursor.execute("""
                SELECT 
                    DATE_FORMAT(o.created_at, '%b %d') as date_label,
                    SUM(oi.subtotal) as revenue
                FROM orders o
                JOIN order_items oi ON o.order_id = oi.order_id
                WHERE o.seller_id = %s
                    AND o.order_status IN ('in_transit', 'delivered')
                    AND o.created_at >= %s
                    AND o.created_at <= %s
                GROUP BY WEEK(o.created_at, 1), date_label
                ORDER BY WEEK(o.created_at, 1)
            """, (seller_id, start_date, end_date))
            revenue_data = cursor.fetchall()
            revenue_label = 'Weekly Revenue'
        else:
            cursor.execute("""
                SELECT 
                    DATE_FORMAT(o.created_at, '%b %Y') as date_label,
                    SUM(oi.subtotal) as revenue
                FROM orders o
                JOIN order_items oi ON o.order_id = oi.order_id
                WHERE o.seller_id = %s
                    AND o.order_status IN ('in_transit', 'delivered')
                    AND o.created_at >= %s
                    AND o.created_at <= %s
                GROUP BY YEAR(o.created_at), MONTH(o.created_at), date_label
                ORDER BY YEAR(o.created_at), MONTH(o.created_at)
            """, (seller_id, start_date, end_date))
            revenue_data = cursor.fetchall()
            revenue_label = 'Monthly Revenue'
        
        close_db_connection(connection, cursor)
        
        return jsonify({
            'units': {
                'labels': [u['date_label'] for u in units_data] if units_data else [],
                'data': [int(u['units_sold']) for u in units_data] if units_data else [],
                'label': units_label
            },
            'revenue': {
                'labels': [r['date_label'] for r in revenue_data] if revenue_data else [],
                'data': [float(r['revenue']) for r in revenue_data] if revenue_data else [],
                'label': revenue_label
            }
        })
        
    except Exception as e:
        print(f"Error getting chart data: {e}")
        return jsonify({'error': str(e)}), 500


@seller_product_sales_bp.route('/seller/product-sales/export-excel')
@seller_required
def export_excel():
    """Export seller product sales to Excel format"""
    try:
        seller_id = session.get('seller_id')
        date_from = request.args.get('dateFrom')
        date_to = request.args.get('dateTo')
        
        print(f"📊 Excel Export Request - Seller ID: {seller_id}, Date From: {date_from}, Date To: {date_to}")
        
        connection = get_db_connection()
        if not connection:
            print("❌ Database connection failed")
            return "Database connection failed", 500
        
        cursor = connection.cursor(dictionary=True)
        
        # Get seller info
        cursor.execute("""
            SELECT shop_name, first_name, last_name
            FROM sellers
            WHERE seller_id = %s
        """, (seller_id,))
        seller_info = cursor.fetchone()
        
        # Calculate date range
        if date_from and date_to:
            start_date = datetime.strptime(date_from, '%Y-%m-%d')
            end_date = datetime.strptime(date_to, '%Y-%m-%d')
            end_date = end_date.replace(hour=23, minute=59, second=59)
        else:
            end_date = datetime.now()
            start_date = end_date - timedelta(days=30)
        
        # Get sales summary
        cursor.execute("""
            SELECT 
                COALESCE(SUM(oi.quantity), 0) as total_products_sold,
                COALESCE(SUM(oi.subtotal), 0) as total_revenue,
                COUNT(DISTINCT o.order_id) as total_orders
            FROM orders o
            JOIN order_items oi ON o.order_id = oi.order_id
            WHERE o.seller_id = %s
                AND o.order_status = 'delivered'
                AND o.order_received = TRUE
                AND o.created_at >= %s
                AND o.created_at <= %s
        """, (seller_id, start_date, end_date))
        summary = cursor.fetchone()
        
        # Get commission
        cursor.execute("""
            SELECT COALESCE(SUM(commission_amount), 0) as total_commission
            FROM orders
            WHERE seller_id = %s
                AND order_status = 'delivered'
                AND order_received = TRUE
                AND created_at >= %s
                AND created_at <= %s
        """, (seller_id, start_date, end_date))
        commission_data = cursor.fetchone()
        
        # Get detailed sales per product
        cursor.execute("""
            SELECT 
                p.product_name,
                SUM(oi.quantity) as units_sold,
                SUM(oi.subtotal) as revenue,
                MIN(o.created_at) as first_order,
                MAX(o.created_at) as latest_order,
                COUNT(DISTINCT o.order_id) as order_count
            FROM order_items oi
            JOIN orders o ON oi.order_id = o.order_id
            JOIN products p ON oi.product_id = p.product_id
            WHERE o.seller_id = %s
                AND o.order_status = 'delivered'
                AND o.order_received = TRUE
                AND o.created_at >= %s
                AND o.created_at <= %s
            GROUP BY oi.product_id, p.product_name
            ORDER BY units_sold DESC
        """, (seller_id, start_date, end_date))
        sales_details = cursor.fetchall()
        
        # Get time-based breakdown (weekly or monthly based on date range)
        days_diff = (end_date - start_date).days
        
        if days_diff <= 30:
            # Weekly breakdown for 30 days or less
            cursor.execute("""
                SELECT 
                    DATE_FORMAT(o.created_at, '%b %d') as period_label,
                    WEEK(o.created_at, 1) as period_num,
                    SUM(oi.quantity) as units_sold,
                    SUM(oi.subtotal) as revenue,
                    COUNT(DISTINCT o.order_id) as order_count
                FROM orders o
                JOIN order_items oi ON o.order_id = oi.order_id
                WHERE o.seller_id = %s
                    AND o.order_status = 'delivered'
                    AND o.order_received = TRUE
                    AND o.created_at >= %s
                    AND o.created_at <= %s
                GROUP BY WEEK(o.created_at, 1), period_label
                ORDER BY WEEK(o.created_at, 1)
            """, (seller_id, start_date, end_date))
            time_breakdown = cursor.fetchall()
            breakdown_type = "Weekly"
        else:
            # Monthly breakdown for longer periods
            cursor.execute("""
                SELECT 
                    DATE_FORMAT(o.created_at, '%b %Y') as period_label,
                    YEAR(o.created_at) as year_num,
                    MONTH(o.created_at) as month_num,
                    SUM(oi.quantity) as units_sold,
                    SUM(oi.subtotal) as revenue,
                    COUNT(DISTINCT o.order_id) as order_count
                FROM orders o
                JOIN order_items oi ON o.order_id = oi.order_id
                WHERE o.seller_id = %s
                    AND o.order_status = 'delivered'
                    AND o.order_received = TRUE
                    AND o.created_at >= %s
                    AND o.created_at <= %s
                GROUP BY YEAR(o.created_at), MONTH(o.created_at), period_label
                ORDER BY YEAR(o.created_at), MONTH(o.created_at)
            """, (seller_id, start_date, end_date))
            time_breakdown = cursor.fetchall()
            breakdown_type = "Monthly"
        
        # Don't close connection yet - we need it for Product Sold data
        # cursor.close()
        # connection.close()
        
        # Create Excel workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Product Sales"
        
        # Define styles
        header_font = Font(name='Calibri', size=18, bold=True, color="1F4E78")
        subheader_font = Font(name='Calibri', size=14, bold=True, color="1F4E78")
        summary_header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        summary_header_font = Font(name='Calibri', size=11, bold=True, color="FFFFFF")
        table_header_fill = PatternFill(start_color="5B9BD5", end_color="5B9BD5", fill_type="solid")
        table_header_font = Font(name='Calibri', size=11, bold=True, color="FFFFFF")
        alt_row_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        
        border = Border(
            left=Side(style='thin', color='000000'),
            right=Side(style='thin', color='000000'),
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000')
        )
        
        center_align = Alignment(horizontal='center', vertical='center')
        left_align = Alignment(horizontal='left', vertical='center')
        right_align = Alignment(horizontal='right', vertical='center')
        
        current_row = 1
        
        # Header
        ws.merge_cells(f'A{current_row}:F{current_row}')
        cell = ws[f'A{current_row}']
        cell.value = seller_info['shop_name'] if seller_info else "Shop"
        cell.font = header_font
        cell.alignment = center_align
        current_row += 1
        
        ws.merge_cells(f'A{current_row}:F{current_row}')
        cell = ws[f'A{current_row}']
        cell.value = "PRODUCT SALES REPORT"
        cell.font = subheader_font
        cell.alignment = center_align
        current_row += 2
        
        # Report info
        ws[f'A{current_row}'] = "Report Generated:"
        ws[f'A{current_row}'].font = Font(bold=True)
        ws[f'B{current_row}'] = datetime.now().strftime('%B %d, %Y at %I:%M %p')
        current_row += 1
        
        ws[f'A{current_row}'] = "Report Period:"
        ws[f'A{current_row}'].font = Font(bold=True)
        ws[f'B{current_row}'] = f"{start_date.strftime('%B %d, %Y')} to {end_date.strftime('%B %d, %Y')}"
        current_row += 2
        
        # Summary section
        ws.merge_cells(f'A{current_row}:B{current_row}')
        cell = ws[f'A{current_row}']
        cell.value = "SALES SUMMARY"
        cell.fill = summary_header_fill
        cell.font = summary_header_font
        cell.alignment = center_align
        cell.border = border
        ws[f'B{current_row}'].border = border
        current_row += 1
        
        # Summary data
        summary_data = [
            ("Total Orders", int(summary['total_orders']) if summary else 0),
            ("Total Products Sold", int(summary['total_products_sold']) if summary else 0),
            ("Total Revenue", f"₱{float(summary['total_revenue']):,.2f}" if summary else "₱0.00"),
            ("Platform Commission", f"₱{float(commission_data['total_commission']):,.2f}" if commission_data else "₱0.00"),
            ("Net Earnings", f"₱{(float(summary['total_revenue']) - float(commission_data['total_commission'])):,.2f}" if summary and commission_data else "₱0.00")
        ]
        
        for label, value in summary_data:
            ws[f'A{current_row}'] = label
            ws[f'A{current_row}'].font = Font(bold=True)
            ws[f'A{current_row}'].border = border
            ws[f'B{current_row}'] = value
            ws[f'B{current_row}'].border = border
            ws[f'B{current_row}'].alignment = right_align
            current_row += 1
        
        current_row += 1
        
        # Product sales table
        ws.merge_cells(f'A{current_row}:F{current_row}')
        cell = ws[f'A{current_row}']
        cell.value = "PRODUCT SALES DETAILS"
        cell.fill = summary_header_fill
        cell.font = summary_header_font
        cell.alignment = center_align
        for col in ['A', 'B', 'C', 'D', 'E', 'F']:
            ws[f'{col}{current_row}'].border = border
        current_row += 1
        
        # Table headers
        headers = ['Product Name', 'Units Sold', 'Revenue', 'Orders', 'First Sale', 'Latest Sale']
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=current_row, column=col_idx)
            cell.value = header
            cell.fill = table_header_fill
            cell.font = table_header_font
            cell.alignment = center_align
            cell.border = border
        current_row += 1
        
        # Table data
        for idx, sale in enumerate(sales_details):
            row_fill = alt_row_fill if idx % 2 == 0 else None
            
            cell = ws.cell(row=current_row, column=1, value=sale['product_name'])
            cell.border = border
            if row_fill:
                cell.fill = row_fill
            cell.alignment = left_align
            
            cell = ws.cell(row=current_row, column=2, value=int(sale['units_sold']))
            cell.border = border
            if row_fill:
                cell.fill = row_fill
            cell.alignment = center_align
            
            cell = ws.cell(row=current_row, column=3, value=f"₱{float(sale['revenue']):,.2f}")
            cell.border = border
            if row_fill:
                cell.fill = row_fill
            cell.alignment = right_align
            
            cell = ws.cell(row=current_row, column=4, value=int(sale['order_count']))
            cell.border = border
            if row_fill:
                cell.fill = row_fill
            cell.alignment = center_align
            
            cell = ws.cell(row=current_row, column=5, value=sale['first_order'].strftime('%b %d, %Y'))
            cell.border = border
            if row_fill:
                cell.fill = row_fill
            cell.alignment = center_align
            
            cell = ws.cell(row=current_row, column=6, value=sale['latest_order'].strftime('%b %d, %Y'))
            cell.border = border
            if row_fill:
                cell.fill = row_fill
            cell.alignment = center_align
            
            current_row += 1
        
        current_row += 2
        
        # ============================================
        # TIME-BASED BREAKDOWN TABLE
        # ============================================
        ws.merge_cells(f'A{current_row}:E{current_row}')
        cell = ws[f'A{current_row}']
        cell.value = f"{breakdown_type.upper()} SALES BREAKDOWN"
        cell.fill = summary_header_fill
        cell.font = summary_header_font
        cell.alignment = center_align
        for col in ['A', 'B', 'C', 'D', 'E']:
            ws[f'{col}{current_row}'].border = border
        current_row += 1
        
        # Time breakdown table headers
        time_headers = ['Period', 'Orders', 'Units Sold', 'Revenue', 'Avg Order Value']
        for col_idx, header in enumerate(time_headers, start=1):
            cell = ws.cell(row=current_row, column=col_idx)
            cell.value = header
            cell.fill = table_header_fill
            cell.font = table_header_font
            cell.alignment = center_align
            cell.border = border
        current_row += 1
        
        # Time breakdown data
        for idx, period in enumerate(time_breakdown):
            row_fill = alt_row_fill if idx % 2 == 0 else None
            avg_order_value = float(period['revenue']) / int(period['order_count']) if period['order_count'] > 0 else 0
            
            cell = ws.cell(row=current_row, column=1, value=period['period_label'])
            cell.border = border
            if row_fill:
                cell.fill = row_fill
            cell.alignment = left_align
            
            cell = ws.cell(row=current_row, column=2, value=int(period['order_count']))
            cell.border = border
            if row_fill:
                cell.fill = row_fill
            cell.alignment = center_align
            
            cell = ws.cell(row=current_row, column=3, value=int(period['units_sold']))
            cell.border = border
            if row_fill:
                cell.fill = row_fill
            cell.alignment = center_align
            
            cell = ws.cell(row=current_row, column=4, value=f"₱{float(period['revenue']):,.2f}")
            cell.border = border
            if row_fill:
                cell.fill = row_fill
            cell.alignment = right_align
            
            cell = ws.cell(row=current_row, column=5, value=f"₱{avg_order_value:,.2f}")
            cell.border = border
            if row_fill:
                cell.fill = row_fill
            cell.alignment = right_align
            
            current_row += 1
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 40
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 18
        ws.column_dimensions['D'].width = 12
        ws.column_dimensions['E'].width = 18
        ws.column_dimensions['F'].width = 18
        
        # ============================================
        # PRODUCT SOLD SHEET (NEW)
        # ============================================
        ws2 = wb.create_sheet(title="Product Sold")
        
        # Get Product Sold data (delivered and confirmed orders)
        # Reuse existing cursor
        cursor.execute("""
            SELECT 
                o.order_id,
                o.order_number,
                o.subtotal as order_total,
                o.commission_amount,
                o.updated_at as order_received_date,
                CONCAT(b.first_name, ' ', b.last_name) as buyer_name
            FROM orders o
            JOIN buyers b ON o.buyer_id = b.buyer_id
            WHERE o.seller_id = %s 
            AND o.order_status = 'delivered'
            AND o.order_received = TRUE
            AND o.created_at >= %s
            AND o.created_at <= %s
            ORDER BY o.updated_at DESC
        """, (seller_id, start_date, end_date))
        sold_orders = cursor.fetchall()
        
        # Get order items for each sold order
        for order in sold_orders:
            cursor.execute("""
                SELECT 
                    product_name,
                    variant_color,
                    variant_size,
                    quantity,
                    unit_price,
                    subtotal
                FROM order_items
                WHERE order_id = %s
            """, (order['order_id'],))
            order['items'] = cursor.fetchall()
        
        # Now close the connection after all queries are done
        cursor.close()
        connection.close()
        
        current_row = 1
        
        # Header
        ws2.merge_cells(f'A{current_row}:F{current_row}')
        cell = ws2[f'A{current_row}']
        cell.value = seller_info['shop_name'] if seller_info else "Shop"
        cell.font = header_font
        cell.alignment = center_align
        current_row += 1
        
        ws2.merge_cells(f'A{current_row}:F{current_row}')
        cell = ws2[f'A{current_row}']
        cell.value = "PRODUCT SOLD REPORT"
        cell.font = subheader_font
        cell.alignment = center_align
        current_row += 2
        
        # Report info
        ws2[f'A{current_row}'] = "Report Generated:"
        ws2[f'A{current_row}'].font = Font(bold=True)
        ws2[f'B{current_row}'] = datetime.now().strftime('%B %d, %Y at %I:%M %p')
        current_row += 1
        
        ws2[f'A{current_row}'] = "Report Period:"
        ws2[f'A{current_row}'].font = Font(bold=True)
        ws2[f'B{current_row}'] = f"{start_date.strftime('%B %d, %Y')} to {end_date.strftime('%B %d, %Y')}"
        current_row += 2
        
        # Product Sold table
        ws2.merge_cells(f'A{current_row}:F{current_row}')
        cell = ws2[f'A{current_row}']
        cell.value = "PRODUCT SOLD DETAILS"
        cell.fill = summary_header_fill
        cell.font = summary_header_font
        cell.alignment = center_align
        for col in ['A', 'B', 'C', 'D', 'E', 'F']:
            ws2[f'{col}{current_row}'].border = border
        current_row += 1
        
        # Table headers
        sold_headers = ['#', 'Products', 'Order Received', 'Buyer', 'Order Total', 'Commission (5%)']
        for col_idx, header in enumerate(sold_headers, start=1):
            cell = ws2.cell(row=current_row, column=col_idx)
            cell.value = header
            cell.fill = table_header_fill
            cell.font = table_header_font
            cell.alignment = center_align
            cell.border = border
        current_row += 1
        
        # Table data
        for idx, order in enumerate(sold_orders, start=1):
            row_fill = alt_row_fill if idx % 2 == 0 else None
            
            # Order number
            cell = ws2.cell(row=current_row, column=1, value=f"#{idx:03d}")
            cell.border = border
            if row_fill:
                cell.fill = row_fill
            cell.alignment = center_align
            
            # Products (with variants and quantities)
            products_text = []
            for item in order['items']:
                variant_parts = []
                if item['variant_color']:
                    variant_parts.append(item['variant_color'])
                if item['variant_size']:
                    variant_parts.append(item['variant_size'])
                variant_text = f" ({', '.join(variant_parts)})" if variant_parts else ""
                products_text.append(f"{item['product_name']}{variant_text} × {item['quantity']}")
            
            cell = ws2.cell(row=current_row, column=2, value="\n".join(products_text))
            cell.border = border
            if row_fill:
                cell.fill = row_fill
            cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
            
            # Order Received Date
            cell = ws2.cell(row=current_row, column=3, value=order['order_received_date'].strftime('%b %d, %Y'))
            cell.border = border
            if row_fill:
                cell.fill = row_fill
            cell.alignment = center_align
            
            # Buyer
            cell = ws2.cell(row=current_row, column=4, value=order['buyer_name'])
            cell.border = border
            if row_fill:
                cell.fill = row_fill
            cell.alignment = left_align
            
            # Order Total
            cell = ws2.cell(row=current_row, column=5, value=f"₱{float(order['order_total']):,.2f}")
            cell.border = border
            if row_fill:
                cell.fill = row_fill
            cell.alignment = right_align
            cell.font = Font(color="008000", bold=True)  # Green for positive
            
            # Commission
            cell = ws2.cell(row=current_row, column=6, value=f"₱{float(order['commission_amount']):,.2f}")
            cell.border = border
            if row_fill:
                cell.fill = row_fill
            cell.alignment = right_align
            cell.font = Font(color="FF0000", bold=True)  # Red for negative
            
            # Set row height for wrapped text
            ws2.row_dimensions[current_row].height = 15 * len(order['items'])
            
            current_row += 1
        
        # Adjust column widths for Product Sold sheet
        ws2.column_dimensions['A'].width = 8
        ws2.column_dimensions['B'].width = 50
        ws2.column_dimensions['C'].width = 18
        ws2.column_dimensions['D'].width = 25
        ws2.column_dimensions['E'].width = 18
        ws2.column_dimensions['F'].width = 18
        
        # Save to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Generate filename
        shop_name_safe = seller_info['shop_name'].replace(' ', '_').replace("'", "")
        filename = f"Product_Sales_{shop_name_safe}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        print(f"✅ Excel file created successfully: {filename}")
        
        # Create response
        response = make_response(output.getvalue())
        response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        response.headers['Content-Disposition'] = f'attachment; filename={filename}'
        
        return response
        
    except Exception as e:
        print(f"❌ Error exporting to Excel: {e}")
        import traceback
        traceback.print_exc()
        return f"Error exporting to Excel: {str(e)}", 500
