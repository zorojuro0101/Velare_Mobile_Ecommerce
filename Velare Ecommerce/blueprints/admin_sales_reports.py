from flask import Blueprint, render_template, make_response, request
from database.db_config import get_db_connection, close_db_connection
import csv
from io import StringIO, BytesIO
from datetime import datetime
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.lib.enums import TA_CENTER, TA_RIGHT
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

admin_sales_reports_bp = Blueprint('admin_sales_reports', __name__)

@admin_sales_reports_bp.route('/admin/sales-reports')
def admin_sales_reports():
    connection = get_db_connection()
    if not connection:
        return render_template('admin/admin_sales_reports.html', error="Database connection failed")
    
    try:
        cursor = connection.cursor(dictionary=True)
        
        # Get total sales from delivered orders
        cursor.execute("""
            SELECT SUM(total_amount) as total 
            FROM orders 
            WHERE order_status = 'delivered'
        """)
        result = cursor.fetchone()
        total_sales = float(result['total'] if result['total'] else 0.0)
        
        # Get total delivered orders
        cursor.execute("""
            SELECT COUNT(*) as count 
            FROM orders 
            WHERE order_status = 'delivered'
        """)
        total_orders = cursor.fetchone()['count']
        
        # Get total commission from delivered orders
        cursor.execute("""
            SELECT SUM(commission_amount) as total 
            FROM orders 
            WHERE order_status = 'delivered'
        """)
        result = cursor.fetchone()
        total_commission = float(result['total'] if result['total'] else 0.0)
        
        # Get sales by category - Show ALL predefined categories even with 0 sales
        # Predefined categories list
        predefined_categories = [
            'Dresses', 'Skirts', 'Tops', 'Blouses', 'Activewear',
            'Yoga Pants', 'Lingerie', 'Sleepwear', 'Jackets', 'Coats',
            'Shoes', 'Accessories'
        ]
        
        cursor.execute("""
            SELECT 
                all_cats.category,
                COALESCE(SUM(oi.subtotal), 0) as total_sales,
                COUNT(DISTINCT CASE WHEN o.order_status = 'delivered' THEN oi.order_id END) as order_count,
                COALESCE(SUM(oi.quantity), 0) as items_sold
            FROM (
                SELECT 'Dresses' as category UNION ALL
                SELECT 'Skirts' UNION ALL
                SELECT 'Tops' UNION ALL
                SELECT 'Blouses' UNION ALL
                SELECT 'Activewear' UNION ALL
                SELECT 'Yoga Pants' UNION ALL
                SELECT 'Lingerie' UNION ALL
                SELECT 'Sleepwear' UNION ALL
                SELECT 'Jackets' UNION ALL
                SELECT 'Coats' UNION ALL
                SELECT 'Shoes' UNION ALL
                SELECT 'Accessories'
            ) all_cats
            LEFT JOIN products p ON all_cats.category = p.category
            LEFT JOIN order_items oi ON p.product_id = oi.product_id
            LEFT JOIN orders o ON oi.order_id = o.order_id AND o.order_status = 'delivered'
            GROUP BY all_cats.category
            ORDER BY total_sales DESC, all_cats.category ASC
        """)
        sales_by_category = cursor.fetchall()
        
        # Format category data
        formatted_categories = []
        for cat in sales_by_category:
            formatted_categories.append({
                'category': cat['category'],
                'total_sales': float(cat['total_sales']),
                'order_count': cat['order_count'],
                'items_sold': cat['items_sold'],
                'percentage': (float(cat['total_sales']) / total_sales * 100) if total_sales > 0 else 0
            })
        
        # Get top sellers by sales
        cursor.execute("""
            SELECT 
                s.seller_id,
                s.shop_name,
                s.first_name,
                s.last_name,
                SUM(o.total_amount) as total_sales,
                COUNT(DISTINCT o.order_id) as order_count,
                AVG(o.total_amount) as avg_order_value
            FROM orders o
            JOIN sellers s ON o.seller_id = s.seller_id
            WHERE o.order_status = 'delivered'
            GROUP BY s.seller_id, s.shop_name, s.first_name, s.last_name
            ORDER BY total_sales DESC
            LIMIT 10
        """)
        top_sellers = cursor.fetchall()
        
        # Format seller data
        formatted_sellers = []
        for seller in top_sellers:
            formatted_sellers.append({
                'seller_id': seller['seller_id'],
                'store_name': seller['shop_name'],
                'name': f"{seller['first_name']} {seller['last_name']}",
                'total_sales': float(seller['total_sales']),
                'order_count': seller['order_count'],
                'avg_order_value': float(seller['avg_order_value'])
            })
        
        # Get sales trend (last 7 days)
        cursor.execute("""
            SELECT 
                DATE(o.created_at) as sale_date,
                SUM(o.total_amount) as daily_sales,
                COUNT(DISTINCT o.order_id) as daily_orders
            FROM orders o
            WHERE o.order_status = 'delivered' 
                AND o.created_at >= DATE_SUB(CURDATE(), INTERVAL 7 DAY)
            GROUP BY DATE(o.created_at)
            ORDER BY sale_date ASC
        """)
        sales_trend = cursor.fetchall()
        
        # Format trend data
        formatted_trend = []
        for trend in sales_trend:
            formatted_trend.append({
                'date': trend['sale_date'].strftime('%b %d'),
                'sales': float(trend['daily_sales']),
                'orders': trend['daily_orders']
            })
        
        # Get detailed sales breakdown (last 20 delivered orders)
        cursor.execute("""
            SELECT 
                o.order_id,
                o.order_number,
                o.total_amount,
                o.commission_amount,
                o.created_at,
                b.first_name as buyer_first,
                b.last_name as buyer_last,
                s.shop_name,
                COUNT(oi.order_item_id) as item_count,
                GROUP_CONCAT(DISTINCT p.category SEPARATOR ', ') as categories
            FROM orders o
            JOIN buyers b ON o.buyer_id = b.buyer_id
            JOIN sellers s ON o.seller_id = s.seller_id
            LEFT JOIN order_items oi ON o.order_id = oi.order_id
            LEFT JOIN products p ON oi.product_id = p.product_id
            WHERE o.order_status = 'delivered'
            GROUP BY o.order_id, o.order_number, o.total_amount, o.commission_amount, 
                     o.created_at, b.first_name, b.last_name, s.shop_name
            ORDER BY o.created_at DESC
            LIMIT 20
        """)
        detailed_sales = cursor.fetchall()
        
        # Format detailed sales data
        formatted_detailed = []
        for sale in detailed_sales:
            formatted_detailed.append({
                'order_id': sale['order_id'],
                'order_number': sale['order_number'],
                'amount': float(sale['total_amount']),
                'commission': float(sale['commission_amount']),
                'date': sale['created_at'].strftime('%b %d, %Y'),
                'buyer': f"{sale['buyer_first']} {sale['buyer_last']}",
                'seller': sale['shop_name'],
                'item_count': sale['item_count'],
                'categories': sale['categories'] or 'N/A'
            })
        
        cursor.close()
        
        stats = {
            'total_sales': total_sales,
            'total_orders': total_orders,
            'total_commission': total_commission,
            'sales_by_category': formatted_categories,
            'top_sellers': formatted_sellers,
            'sales_trend': formatted_trend,
            'detailed_sales': formatted_detailed
        }
        
    except Exception as e:
        print(f"Error fetching sales reports: {e}")
        stats = {
            'total_sales': 0.0,
            'total_orders': 0,
            'total_commission': 0.0,
            'sales_by_category': [],
            'top_sellers': [],
            'sales_trend': [],
            'detailed_sales': []
        }
        if 'cursor' in locals():
            cursor.close()
    finally:
        close_db_connection(connection)
    
    return render_template('admin/admin_sales_reports.html', stats=stats)


@admin_sales_reports_bp.route('/admin/sales-reports/export-csv')
def export_csv():
    """Export sales reports to CSV format"""
    connection = get_db_connection()
    if not connection:
        return "Database connection failed", 500
    
    try:
        cursor = connection.cursor(dictionary=True)
        
        # Get filter parameters
        start_date = request.args.get('start_date', '')
        end_date = request.args.get('end_date', '')
        category = request.args.get('category', '')
        
        # Build query with filters
        params = []
        
        if category:
            # If filtering by category, use a subquery to get orders with that category
            query = """
                SELECT 
                    o.order_number,
                    o.created_at as order_date,
                    s.shop_name as seller_name,
                    CONCAT(b.first_name, ' ', b.last_name) as buyer_name,
                    COUNT(oi.order_item_id) as total_items,
                    o.total_amount,
                    o.commission_amount,
                    o.shipping_fee,
                    GROUP_CONCAT(DISTINCT p.category SEPARATOR ', ') as categories,
                    GROUP_CONCAT(CONCAT(p.product_name, ' (', oi.quantity, 'x)') SEPARATOR '; ') as products
                FROM orders o
                JOIN buyers b ON o.buyer_id = b.buyer_id
                JOIN sellers s ON o.seller_id = s.seller_id
                LEFT JOIN order_items oi ON o.order_id = oi.order_id
                LEFT JOIN products p ON oi.product_id = p.product_id
                WHERE o.order_status = 'delivered'
                    AND o.order_id IN (
                        SELECT DISTINCT oi2.order_id 
                        FROM order_items oi2 
                        JOIN products p2 ON oi2.product_id = p2.product_id 
                        WHERE p2.category = %s
                    )
            """
            params.append(category)
        else:
            query = """
                SELECT 
                    o.order_number,
                    o.created_at as order_date,
                    s.shop_name as seller_name,
                    CONCAT(b.first_name, ' ', b.last_name) as buyer_name,
                    COUNT(oi.order_item_id) as total_items,
                    o.total_amount,
                    o.commission_amount,
                    o.shipping_fee,
                    GROUP_CONCAT(DISTINCT p.category SEPARATOR ', ') as categories,
                    GROUP_CONCAT(CONCAT(p.product_name, ' (', oi.quantity, 'x)') SEPARATOR '; ') as products
                FROM orders o
                JOIN buyers b ON o.buyer_id = b.buyer_id
                JOIN sellers s ON o.seller_id = s.seller_id
                LEFT JOIN order_items oi ON o.order_id = oi.order_id
                LEFT JOIN products p ON oi.product_id = p.product_id
                WHERE o.order_status = 'delivered'
            """
        
        if start_date:
            query += " AND DATE(o.created_at) >= %s"
            params.append(start_date)
        if end_date:
            query += " AND DATE(o.created_at) <= %s"
            params.append(end_date)
            
        query += """
            GROUP BY o.order_id, o.order_number, o.created_at, s.shop_name, 
                     b.first_name, b.last_name, o.total_amount, o.commission_amount, o.shipping_fee
            ORDER BY o.created_at DESC
        """
        
        cursor.execute(query, params)
        sales_data = cursor.fetchall()
        
        # Get summary statistics
        summary_query = """
            SELECT 
                COUNT(*) as total_orders,
                SUM(total_amount) as total_sales,
                SUM(commission_amount) as total_commission,
                AVG(total_amount) as avg_order_value
            FROM orders 
            WHERE order_status = 'delivered'
        """
        summary_params = []
        if start_date:
            summary_query += " AND DATE(created_at) >= %s"
            summary_params.append(start_date)
        if end_date:
            summary_query += " AND DATE(created_at) <= %s"
            summary_params.append(end_date)
        
        cursor.execute(summary_query, summary_params)
        summary = cursor.fetchone()
        cursor.close()
        
        # Create CSV with proper business formatting
        output = StringIO()
        writer = csv.writer(output)
        
        # ============================================
        # REPORT HEADER SECTION
        # ============================================
        writer.writerow(['VELÁRE E-COMMERCE PLATFORM'])
        writer.writerow(['SALES REPORT'])
        writer.writerow([])
        writer.writerow(['Report Generated:', datetime.now().strftime('%B %d, %Y at %I:%M %p')])
        
        if start_date or end_date:
            date_range = f"{start_date or 'Beginning'} to {end_date or 'Present'}"
            writer.writerow(['Report Period:', date_range])
        else:
            writer.writerow(['Report Period:', 'All Time'])
            
        if category:
            writer.writerow(['Category Filter:', category])
        
        writer.writerow([])
        writer.writerow(['=' * 80])
        writer.writerow([])
        
        # ============================================
        # EXECUTIVE SUMMARY SECTION
        # ============================================
        writer.writerow(['EXECUTIVE SUMMARY'])
        writer.writerow([])
        
        total_orders = summary['total_orders'] or 0
        total_sales = float(summary['total_sales']) if summary['total_sales'] else 0.0
        total_commission = float(summary['total_commission']) if summary['total_commission'] else 0.0
        avg_order = float(summary['avg_order_value']) if summary['avg_order_value'] else 0.0
        
        writer.writerow(['Metric', 'Value'])
        writer.writerow(['Total Delivered Orders', f"{total_orders:,}"])
        writer.writerow(['Gross Sales Revenue', f"₱{total_sales:,.2f}"])
        writer.writerow(['Platform Commission Earned', f"₱{total_commission:,.2f}"])
        writer.writerow(['Average Order Value', f"₱{avg_order:,.2f}"])
        
        if total_sales > 0:
            commission_rate = (total_commission / total_sales) * 100
            writer.writerow(['Commission Rate', f"{commission_rate:.1f}%"])
        
        writer.writerow([])
        writer.writerow(['=' * 80])
        writer.writerow([])
        
        # ============================================
        # DETAILED TRANSACTION RECORDS
        # ============================================
        writer.writerow(['DETAILED TRANSACTION RECORDS'])
        writer.writerow([])
        writer.writerow([
            'Order No.',
            'Transaction Date',
            'Transaction Time',
            'Seller Name',
            'Buyer Name',
            'Items',
            'Subtotal',
            'Shipping',
            'Total Amount',
            'Commission (5%)',
            'Net to Seller',
            'Category',
            'Products Ordered'
        ])
        
        # Write data rows with calculations
        for row in sales_data:
            # Format datetime
            order_date = row['order_date']
            if hasattr(order_date, 'strftime'):
                date_str = order_date.strftime('%Y-%m-%d')
                time_str = order_date.strftime('%I:%M %p')
            else:
                date_str = str(order_date).split()[0] if order_date else 'N/A'
                time_str = str(order_date).split()[1] if order_date and len(str(order_date).split()) > 1 else 'N/A'
            
            # Calculate values
            total_amount = float(row['total_amount'])
            commission = float(row['commission_amount'])
            shipping = float(row['shipping_fee']) if row['shipping_fee'] else 0.0
            subtotal = total_amount - shipping
            net_to_seller = total_amount - commission
            
            writer.writerow([
                row['order_number'],
                date_str,
                time_str,
                row['seller_name'],
                row['buyer_name'],
                row['total_items'],
                f"₱{subtotal:,.2f}",
                f"₱{shipping:,.2f}",
                f"₱{total_amount:,.2f}",
                f"₱{commission:,.2f}",
                f"₱{net_to_seller:,.2f}",
                (row['categories'] or 'N/A').title(),
                row['products'] or 'N/A'
            ])
        
        # ============================================
        # FOOTER SECTION
        # ============================================
        writer.writerow([])
        writer.writerow(['=' * 80])
        writer.writerow([])
        writer.writerow(['END OF REPORT'])
        writer.writerow(['Total Records:', len(sales_data)])
        writer.writerow([])
        writer.writerow(['Note: All amounts are in Philippine Peso (₱)'])
        writer.writerow(['Commission rate: 5% of total order amount'])
        writer.writerow(['This is a system-generated report and does not require signature'])
        
        # Create response
        output.seek(0)
        response = make_response(output.getvalue())
        response.headers['Content-Type'] = 'text/csv'
        response.headers['Content-Disposition'] = f'attachment; filename=sales_report_{datetime.now().strftime("%Y%m%d_%H%M%S")}.csv'
        
        return response
        
    except Exception as e:
        print(f"Error exporting CSV: {e}")
        return f"Error exporting CSV: {str(e)}", 500
    finally:
        close_db_connection(connection)


@admin_sales_reports_bp.route('/admin/sales-reports/export-pdf')
def export_pdf():
    """Export sales reports to PDF format"""
    connection = get_db_connection()
    if not connection:
        return "Database connection failed", 500
    
    try:
        cursor = connection.cursor(dictionary=True)
        
        # Get filter parameters
        start_date = request.args.get('start_date', '')
        end_date = request.args.get('end_date', '')
        category = request.args.get('category', '')
        
        # Build query with filters
        params = []
        
        if category:
            # If filtering by category, use a subquery to get orders with that category
            query = """
                SELECT 
                    o.order_number,
                    DATE(o.created_at) as order_date,
                    s.shop_name as seller_name,
                    CONCAT(b.first_name, ' ', b.last_name) as buyer_name,
                    COUNT(oi.order_item_id) as total_items,
                    o.total_amount,
                    o.commission_amount
                FROM orders o
                JOIN buyers b ON o.buyer_id = b.buyer_id
                JOIN sellers s ON o.seller_id = s.seller_id
                LEFT JOIN order_items oi ON o.order_id = oi.order_id
                LEFT JOIN products p ON oi.product_id = p.product_id
                WHERE o.order_status = 'delivered'
                    AND o.order_id IN (
                        SELECT DISTINCT oi2.order_id 
                        FROM order_items oi2 
                        JOIN products p2 ON oi2.product_id = p2.product_id 
                        WHERE p2.category = %s
                    )
            """
            params.append(category)
        else:
            query = """
                SELECT 
                    o.order_number,
                    DATE(o.created_at) as order_date,
                    s.shop_name as seller_name,
                    CONCAT(b.first_name, ' ', b.last_name) as buyer_name,
                    COUNT(oi.order_item_id) as total_items,
                    o.total_amount,
                    o.commission_amount
                FROM orders o
                JOIN buyers b ON o.buyer_id = b.buyer_id
                JOIN sellers s ON o.seller_id = s.seller_id
                LEFT JOIN order_items oi ON o.order_id = oi.order_id
                LEFT JOIN products p ON oi.product_id = p.product_id
                WHERE o.order_status = 'delivered'
            """
        
        if start_date:
            query += " AND DATE(o.created_at) >= %s"
            params.append(start_date)
        if end_date:
            query += " AND DATE(o.created_at) <= %s"
            params.append(end_date)
            
        query += """
            GROUP BY o.order_id, o.order_number, o.created_at, s.shop_name, 
                     b.first_name, b.last_name, o.total_amount, o.commission_amount
            ORDER BY o.created_at DESC
        """
        
        cursor.execute(query, params)
        sales_data = cursor.fetchall()
        
        # Get summary statistics
        summary_query = """
            SELECT 
                COUNT(*) as total_orders,
                SUM(total_amount) as total_sales,
                SUM(commission_amount) as total_commission,
                AVG(total_amount) as avg_order_value
            FROM orders 
            WHERE order_status = 'delivered'
        """
        summary_params = []
        if start_date:
            summary_query += " AND DATE(created_at) >= %s"
            summary_params.append(start_date)
        if end_date:
            summary_query += " AND DATE(created_at) <= %s"
            summary_params.append(end_date)
        
        cursor.execute(summary_query, summary_params)
        summary = cursor.fetchone()
        
        # Get category breakdown - Show ALL predefined categories even with 0 sales
        category_query = """
            SELECT 
                all_cats.category,
                COALESCE(SUM(oi.subtotal), 0) as total_sales,
                COUNT(DISTINCT oi.order_id) as order_count
            FROM (
                SELECT 'Dresses' as category UNION ALL
                SELECT 'Skirts' UNION ALL
                SELECT 'Tops' UNION ALL
                SELECT 'Blouses' UNION ALL
                SELECT 'Activewear' UNION ALL
                SELECT 'Yoga Pants' UNION ALL
                SELECT 'Lingerie' UNION ALL
                SELECT 'Sleepwear' UNION ALL
                SELECT 'Jackets' UNION ALL
                SELECT 'Coats' UNION ALL
                SELECT 'Shoes' UNION ALL
                SELECT 'Accessories'
            ) all_cats
            LEFT JOIN products p ON all_cats.category = p.category
            LEFT JOIN order_items oi ON p.product_id = oi.product_id
            LEFT JOIN orders o ON oi.order_id = o.order_id 
                AND o.order_status = 'delivered'
        """
        category_params = []
        
        # Add date filters
        if start_date:
            category_query += " AND DATE(o.created_at) >= %s"
            category_params.append(start_date)
        if end_date:
            category_query += " AND DATE(o.created_at) <= %s"
            category_params.append(end_date)
            
        category_query += """
            GROUP BY all_cats.category
            ORDER BY total_sales DESC, all_cats.category ASC
        """
        
        cursor.execute(category_query, category_params)
        category_data = cursor.fetchall()
        
        # Get rider earnings data
        rider_query = """
            SELECT 
                r.rider_id,
                CONCAT(r.first_name, ' ', r.last_name) as rider_name,
                r.phone_number,
                COUNT(DISTINCT d.delivery_id) as total_deliveries,
                SUM(d.delivery_fee) as total_delivery_fees,
                SUM(d.rider_earnings) as total_earnings,
                AVG(d.rider_earnings) as avg_earnings_per_delivery
            FROM deliveries d
            JOIN riders r ON d.rider_id = r.rider_id
            JOIN orders o ON d.order_id = o.order_id
            WHERE d.status = 'delivered'
                AND o.order_status = 'delivered'
        """
        rider_params = []
        if start_date:
            rider_query += " AND DATE(d.delivered_at) >= %s"
            rider_params.append(start_date)
        if end_date:
            rider_query += " AND DATE(d.delivered_at) <= %s"
            rider_params.append(end_date)
        rider_query += """
            GROUP BY r.rider_id, r.first_name, r.last_name, r.phone_number
            ORDER BY total_earnings DESC
        """
        
        cursor.execute(rider_query, rider_params)
        rider_data = cursor.fetchall()
        cursor.close()
        
        # Create PDF
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=0.5*inch, bottomMargin=0.5*inch)
        elements = []
        
        # Styles
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=24,
            textColor=colors.HexColor('#2c3e50'),
            spaceAfter=30,
            alignment=TA_CENTER
        )
        heading_style = ParagraphStyle(
            'CustomHeading',
            parent=styles['Heading2'],
            fontSize=14,
            textColor=colors.HexColor('#34495e'),
            spaceAfter=12,
            spaceBefore=12
        )
        
        # Title
        elements.append(Paragraph("Veláre Sales Report", title_style))
        elements.append(Spacer(1, 0.2*inch))
        
        # Report Info
        info_data = [
            ['Generated:', datetime.now().strftime('%Y-%m-%d %H:%M:%S')],
        ]
        if start_date or end_date:
            info_data.append(['Date Range:', f"{start_date or 'Start'} to {end_date or 'Present'}"])
        if category:
            info_data.append(['Category Filter:', category])
            
        info_table = Table(info_data, colWidths=[1.5*inch, 4*inch])
        info_table.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('TEXTCOLOR', (0, 0), (0, -1), colors.HexColor('#7f8c8d')),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ]))
        elements.append(info_table)
        elements.append(Spacer(1, 0.3*inch))
        
        # Summary Statistics
        elements.append(Paragraph("Summary Statistics", heading_style))
        summary_data = [
            ['Metric', 'Value'],
            ['Total Orders', f"{summary['total_orders']:,}"],
            ['Total Sales', f"PHP {summary['total_sales']:,.2f}" if summary['total_sales'] else 'PHP 0.00'],
            ['Total Commission', f"PHP {summary['total_commission']:,.2f}" if summary['total_commission'] else 'PHP 0.00'],
            ['Average Order Value', f"PHP {summary['avg_order_value']:,.2f}" if summary['avg_order_value'] else 'PHP 0.00'],
        ]
        
        summary_table = Table(summary_data, colWidths=[3*inch, 2.5*inch])
        summary_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#3498db')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.grey),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 10),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f8f9fa')]),
        ]))
        elements.append(summary_table)
        elements.append(Spacer(1, 0.3*inch))
        
        # Category Breakdown - Always show even if no data
        elements.append(Paragraph("Sales by Category", heading_style))
        cat_data = [['Category', 'Total Sales', 'Orders']]
        
        if category_data:
            for cat in category_data:
                cat_data.append([
                    cat['category'],
                    f"PHP {cat['total_sales']:,.2f}",
                    str(cat['order_count'])
                ])
        else:
            # Show "No data" row if empty
            cat_data.append(['No category data available', 'PHP 0.00', '0'])
        
        cat_table = Table(cat_data, colWidths=[2*inch, 2*inch, 1.5*inch])
        cat_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2ecc71')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('ALIGN', (1, 0), (-1, -1), 'RIGHT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 11),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('GRID', (0, 0), (-1, -1), 1, colors.grey),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 9),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f8f9fa')]),
        ]))
        elements.append(cat_table)
        elements.append(Spacer(1, 0.3*inch))
        
        # Rider Earnings Breakdown - Always show even if no data
        elements.append(Paragraph("Rider Earnings Summary", heading_style))
        rider_table_data = [['Rider Name', 'Phone', 'Deliveries', 'Total Fees', 'Total Earnings', 'Avg/Delivery']]
        
        if rider_data:
            for rider in rider_data:
                rider_table_data.append([
                    rider['rider_name'],
                    rider['phone_number'] or 'N/A',
                    str(rider['total_deliveries']),
                    f"PHP {rider['total_delivery_fees']:,.2f}",
                    f"PHP {rider['total_earnings']:,.2f}",
                    f"PHP {rider['avg_earnings_per_delivery']:,.2f}"
                ])
        else:
            # Show "No data" row if empty
            rider_table_data.append(['No rider earnings data available', 'N/A', '0', 'PHP 0.00', 'PHP 0.00', 'PHP 0.00'])
        
        rider_table = Table(rider_table_data, colWidths=[1.5*inch, 1.2*inch, 0.9*inch, 1*inch, 1.1*inch, 1*inch])
        rider_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#9b59b6')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('ALIGN', (2, 0), (-1, -1), 'RIGHT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('GRID', (0, 0), (-1, -1), 1, colors.grey),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f8f9fa')]),
        ]))
        elements.append(rider_table)
        elements.append(Spacer(1, 0.3*inch))
        
        # Detailed Sales Data
        elements.append(PageBreak())
        elements.append(Paragraph("Detailed Sales Data", heading_style))
        
        # Table header
        table_data = [['Order #', 'Date', 'Seller', 'Buyer', 'Items', 'Amount', 'Commission']]
        
        # Add data rows (limit to prevent huge PDFs)
        for row in sales_data[:100]:  # Limit to 100 rows
            # Format datetime if it's a datetime object
            order_date = row['order_date']
            if hasattr(order_date, 'strftime'):
                order_date = order_date.strftime('%Y-%m-%d')
            
            table_data.append([
                row['order_number'],
                str(order_date),
                row['seller_name'][:20],  # Truncate long names
                row['buyer_name'][:20],
                str(row['total_items']),
                f"PHP {row['total_amount']:,.2f}",
                f"PHP {row['commission_amount']:,.2f}"
            ])
        
        if len(sales_data) > 100:
            table_data.append(['...', f'Showing 100 of {len(sales_data)} orders', '', '', '', '', ''])
        
        # Create table
        detail_table = Table(table_data, colWidths=[0.8*inch, 0.9*inch, 1.3*inch, 1.3*inch, 0.6*inch, 1*inch, 1*inch])
        detail_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#e74c3c')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('ALIGN', (4, 0), (4, -1), 'CENTER'),
            ('ALIGN', (5, 0), (-1, -1), 'RIGHT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 9),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f8f9fa')]),
        ]))
        elements.append(detail_table)
        
        # Build PDF
        doc.build(elements)
        
        # Create response
        buffer.seek(0)
        response = make_response(buffer.getvalue())
        response.headers['Content-Type'] = 'application/pdf'
        response.headers['Content-Disposition'] = f'attachment; filename=sales_report_{datetime.now().strftime("%Y%m%d_%H%M%S")}.pdf'
        
        return response
        
    except Exception as e:
        print(f"Error exporting PDF: {e}")
        return f"Error exporting PDF: {str(e)}", 500
    finally:
        close_db_connection(connection)



@admin_sales_reports_bp.route('/admin/sales-reports/export-excel')
def export_excel():
    """Export sales reports to Excel format with colors and formatting"""
    connection = get_db_connection()
    if not connection:
        return "Database connection failed", 500
    
    try:
        cursor = connection.cursor(dictionary=True)
        
        # Get filter parameters
        start_date = request.args.get('start_date', '')
        end_date = request.args.get('end_date', '')
        category = request.args.get('category', '')
        
        # Build query with filters
        params = []
        
        if category:
            query = """
                SELECT 
                    o.order_number,
                    o.created_at as order_date,
                    s.shop_name as seller_name,
                    CONCAT(b.first_name, ' ', b.last_name) as buyer_name,
                    COUNT(oi.order_item_id) as total_items,
                    o.total_amount,
                    o.commission_amount,
                    o.shipping_fee,
                    GROUP_CONCAT(DISTINCT p.category SEPARATOR ', ') as categories,
                    GROUP_CONCAT(CONCAT(p.product_name, ' (', oi.quantity, 'x)') SEPARATOR '; ') as products
                FROM orders o
                JOIN buyers b ON o.buyer_id = b.buyer_id
                JOIN sellers s ON o.seller_id = s.seller_id
                LEFT JOIN order_items oi ON o.order_id = oi.order_id
                LEFT JOIN products p ON oi.product_id = p.product_id
                WHERE o.order_status = 'delivered'
                    AND o.order_id IN (
                        SELECT DISTINCT oi2.order_id 
                        FROM order_items oi2 
                        JOIN products p2 ON oi2.product_id = p2.product_id 
                        WHERE p2.category = %s
                    )
            """
            params.append(category)
        else:
            query = """
                SELECT 
                    o.order_number,
                    o.created_at as order_date,
                    s.shop_name as seller_name,
                    CONCAT(b.first_name, ' ', b.last_name) as buyer_name,
                    COUNT(oi.order_item_id) as total_items,
                    o.total_amount,
                    o.commission_amount,
                    o.shipping_fee,
                    GROUP_CONCAT(DISTINCT p.category SEPARATOR ', ') as categories,
                    GROUP_CONCAT(CONCAT(p.product_name, ' (', oi.quantity, 'x)') SEPARATOR '; ') as products
                FROM orders o
                JOIN buyers b ON o.buyer_id = b.buyer_id
                JOIN sellers s ON o.seller_id = s.seller_id
                LEFT JOIN order_items oi ON o.order_id = oi.order_id
                LEFT JOIN products p ON oi.product_id = p.product_id
                WHERE o.order_status = 'delivered'
            """
        
        if start_date:
            query += " AND DATE(o.created_at) >= %s"
            params.append(start_date)
        if end_date:
            query += " AND DATE(o.created_at) <= %s"
            params.append(end_date)
            
        query += """
            GROUP BY o.order_id, o.order_number, o.created_at, s.shop_name, 
                     b.first_name, b.last_name, o.total_amount, o.commission_amount, o.shipping_fee
            ORDER BY o.created_at DESC
        """
        
        cursor.execute(query, params)
        sales_data = cursor.fetchall()
        
        # Get summary statistics
        summary_query = """
            SELECT 
                COUNT(*) as total_orders,
                SUM(total_amount) as total_sales,
                SUM(commission_amount) as total_commission,
                AVG(total_amount) as avg_order_value
            FROM orders 
            WHERE order_status = 'delivered'
        """
        summary_params = []
        if start_date:
            summary_query += " AND DATE(created_at) >= %s"
            summary_params.append(start_date)
        if end_date:
            summary_query += " AND DATE(created_at) <= %s"
            summary_params.append(end_date)
        
        cursor.execute(summary_query, summary_params)
        summary = cursor.fetchone()
        
        # Get rider earnings data
        rider_query = """
            SELECT 
                r.rider_id,
                CONCAT(r.first_name, ' ', r.last_name) as rider_name,
                r.phone_number,
                COUNT(DISTINCT d.delivery_id) as total_deliveries,
                SUM(d.delivery_fee) as total_delivery_fees,
                SUM(d.rider_earnings) as total_earnings,
                AVG(d.rider_earnings) as avg_earnings_per_delivery
            FROM deliveries d
            JOIN riders r ON d.rider_id = r.rider_id
            JOIN orders o ON d.order_id = o.order_id
            WHERE d.status = 'delivered'
                AND o.order_status = 'delivered'
        """
        rider_params = []
        if start_date:
            rider_query += " AND DATE(d.delivered_at) >= %s"
            rider_params.append(start_date)
        if end_date:
            rider_query += " AND DATE(d.delivered_at) <= %s"
            rider_params.append(end_date)
        rider_query += """
            GROUP BY r.rider_id, r.first_name, r.last_name, r.phone_number
            ORDER BY total_earnings DESC
        """
        
        cursor.execute(rider_query, rider_params)
        rider_data = cursor.fetchall()
        cursor.close()
        
        # Create Excel workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Sales Report"
        
        # Define styles
        header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        header_font = Font(name='Calibri', size=14, bold=True, color="FFFFFF")
        
        summary_header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        summary_header_font = Font(name='Calibri', size=11, bold=True, color="FFFFFF")
        
        table_header_fill = PatternFill(start_color="5B9BD5", end_color="5B9BD5", fill_type="solid")
        table_header_font = Font(name='Calibri', size=11, bold=True, color="FFFFFF")
        
        alt_row_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        
        border = Border(
            left=Side(style='thin', color='000000'),
            right=Side(style='thin', color='000000'),
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000')
        )
        
        center_align = Alignment(horizontal='center', vertical='center')
        left_align = Alignment(horizontal='left', vertical='center')
        right_align = Alignment(horizontal='right', vertical='center')
        
        # Row counter
        current_row = 1
        
        # ============================================
        # HEADER SECTION
        # ============================================
        ws.merge_cells(f'A{current_row}:M{current_row}')
        cell = ws[f'A{current_row}']
        cell.value = "VELÁRE E-COMMERCE PLATFORM"
        cell.font = Font(name='Calibri', size=18, bold=True, color="1F4E78")
        cell.alignment = center_align
        current_row += 1
        
        ws.merge_cells(f'A{current_row}:M{current_row}')
        cell = ws[f'A{current_row}']
        cell.value = "SALES REPORT"
        cell.font = Font(name='Calibri', size=16, bold=True, color="1F4E78")
        cell.alignment = center_align
        current_row += 2
        
        # Report info
        ws[f'A{current_row}'] = "Report Generated:"
        ws[f'A{current_row}'].font = Font(bold=True)
        ws[f'B{current_row}'] = datetime.now().strftime('%B %d, %Y at %I:%M %p')
        current_row += 1
        
        ws[f'A{current_row}'] = "Report Period:"
        ws[f'A{current_row}'].font = Font(bold=True)
        if start_date or end_date:
            ws[f'B{current_row}'] = f"{start_date or 'Beginning'} to {end_date or 'Present'}"
        else:
            ws[f'B{current_row}'] = "All Time"
        current_row += 1
        
        if category:
            ws[f'A{current_row}'] = "Category Filter:"
            ws[f'A{current_row}'].font = Font(bold=True)
            ws[f'B{current_row}'] = category
            current_row += 1
        
        current_row += 1
        
        # ============================================
        # EXECUTIVE SUMMARY SECTION
        # ============================================
        ws.merge_cells(f'A{current_row}:B{current_row}')
        cell = ws[f'A{current_row}']
        cell.value = "EXECUTIVE SUMMARY"
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        current_row += 1
        
        # Summary table headers
        ws[f'A{current_row}'] = "Metric"
        ws[f'B{current_row}'] = "Value"
        for col in ['A', 'B']:
            cell = ws[f'{col}{current_row}']
            cell.font = summary_header_font
            cell.fill = summary_header_fill
            cell.alignment = center_align
            cell.border = border
        current_row += 1
        
        # Summary data
        total_orders = summary['total_orders'] or 0
        total_sales = float(summary['total_sales']) if summary['total_sales'] else 0.0
        total_commission = float(summary['total_commission']) if summary['total_commission'] else 0.0
        avg_order = float(summary['avg_order_value']) if summary['avg_order_value'] else 0.0
        
        summary_data = [
            ("Total Delivered Orders", f"{total_orders:,}"),
            ("Gross Sales Revenue", f"₱{total_sales:,.2f}"),
            ("Platform Commission Earned", f"₱{total_commission:,.2f}"),
            ("Average Order Value", f"₱{avg_order:,.2f}"),
        ]
        
        if total_sales > 0:
            commission_rate = (total_commission / total_sales) * 100
            summary_data.append(("Commission Rate", f"{commission_rate:.1f}%"))
        
        for metric, value in summary_data:
            ws[f'A{current_row}'] = metric
            ws[f'B{current_row}'] = value
            ws[f'A{current_row}'].border = border
            ws[f'B{current_row}'].border = border
            ws[f'B{current_row}'].alignment = right_align
            current_row += 1
        
        current_row += 2
        
        # ============================================
        # RIDER EARNINGS SUMMARY
        # ============================================
        if rider_data:
            ws.merge_cells(f'A{current_row}:G{current_row}')
            cell = ws[f'A{current_row}']
            cell.value = "RIDER EARNINGS SUMMARY"
            cell.font = header_font
            cell.fill = PatternFill(start_color="9B59B6", end_color="9B59B6", fill_type="solid")
            cell.alignment = center_align
            current_row += 1
            
            # Rider table headers
            rider_headers = ['Rider Name', 'Phone Number', 'Total Deliveries', 'Total Delivery Fees', 'Total Earnings', 'Avg Earnings/Delivery']
            for col_num, header in enumerate(rider_headers, 1):
                cell = ws.cell(row=current_row, column=col_num)
                cell.value = header
                cell.font = Font(name='Calibri', size=11, bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="A569BD", end_color="A569BD", fill_type="solid")
                cell.alignment = center_align
                cell.border = border
            
            current_row += 1
            
            # Rider data rows
            for idx, rider in enumerate(rider_data):
                rider_row = [
                    rider['rider_name'],
                    rider['phone_number'] or 'N/A',
                    rider['total_deliveries'],
                    f"₱{rider['total_delivery_fees']:,.2f}",
                    f"₱{rider['total_earnings']:,.2f}",
                    f"₱{rider['avg_earnings_per_delivery']:,.2f}"
                ]
                
                for col_num, value in enumerate(rider_row, 1):
                    cell = ws.cell(row=current_row, column=col_num)
                    cell.value = value
                    cell.border = border
                    
                    # Alignment
                    if col_num <= 2:
                        cell.alignment = left_align
                    else:
                        cell.alignment = right_align
                    
                    # Alternate row colors
                    if idx % 2 == 1:
                        cell.fill = alt_row_fill
                
                current_row += 1
            
            # Total row for rider earnings
            ws[f'A{current_row}'] = "TOTAL"
            ws[f'A{current_row}'].font = Font(bold=True)
            ws[f'A{current_row}'].border = border
            ws[f'A{current_row}'].alignment = left_align
            
            total_deliveries = sum(r['total_deliveries'] for r in rider_data)
            total_fees = sum(float(r['total_delivery_fees']) for r in rider_data)
            total_earnings = sum(float(r['total_earnings']) for r in rider_data)
            
            ws[f'B{current_row}'] = ""
            ws[f'B{current_row}'].border = border
            ws[f'C{current_row}'] = total_deliveries
            ws[f'C{current_row}'].font = Font(bold=True)
            ws[f'C{current_row}'].border = border
            ws[f'C{current_row}'].alignment = right_align
            ws[f'D{current_row}'] = f"₱{total_fees:,.2f}"
            ws[f'D{current_row}'].font = Font(bold=True)
            ws[f'D{current_row}'].border = border
            ws[f'D{current_row}'].alignment = right_align
            ws[f'E{current_row}'] = f"₱{total_earnings:,.2f}"
            ws[f'E{current_row}'].font = Font(bold=True)
            ws[f'E{current_row}'].border = border
            ws[f'E{current_row}'].alignment = right_align
            ws[f'F{current_row}'] = ""
            ws[f'F{current_row}'].border = border
            
            current_row += 2
        
        # ============================================
        # DETAILED TRANSACTION RECORDS
        # ============================================
        ws.merge_cells(f'A{current_row}:M{current_row}')
        cell = ws[f'A{current_row}']
        cell.value = "DETAILED TRANSACTION RECORDS"
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        current_row += 1
        
        # Table headers
        headers = [
            "Order No.", "Date", "Time", "Seller", "Buyer", 
            "Items", "Subtotal", "Shipping", "Total", 
            "Commission", "Net to Seller", "Category", "Products"
        ]
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=current_row, column=col_num)
            cell.value = header
            cell.font = table_header_font
            cell.fill = table_header_fill
            cell.alignment = center_align
            cell.border = border
        
        current_row += 1
        
        # Data rows
        for idx, row in enumerate(sales_data):
            # Format datetime
            order_date = row['order_date']
            if hasattr(order_date, 'strftime'):
                date_str = order_date.strftime('%Y-%m-%d')
                time_str = order_date.strftime('%I:%M %p')
            else:
                date_str = str(order_date).split()[0] if order_date else 'N/A'
                time_str = 'N/A'
            
            # Calculate values
            total_amount = float(row['total_amount'])
            commission = float(row['commission_amount'])
            shipping = float(row['shipping_fee']) if row['shipping_fee'] else 0.0
            subtotal = total_amount - shipping
            net_to_seller = total_amount - commission
            
            # Write data
            data_row = [
                row['order_number'],
                date_str,
                time_str,
                row['seller_name'],
                row['buyer_name'],
                row['total_items'],
                f"₱{subtotal:,.2f}",
                f"₱{shipping:,.2f}",
                f"₱{total_amount:,.2f}",
                f"₱{commission:,.2f}",
                f"₱{net_to_seller:,.2f}",
                (row['categories'] or 'N/A').title(),
                row['products'] or 'N/A'
            ]
            
            for col_num, value in enumerate(data_row, 1):
                cell = ws.cell(row=current_row, column=col_num)
                cell.value = value
                cell.border = border
                
                # Alignment
                if col_num <= 5:
                    cell.alignment = left_align
                elif col_num <= 11:
                    cell.alignment = right_align
                else:
                    cell.alignment = left_align
                
                # Alternate row colors
                if idx % 2 == 1:
                    cell.fill = alt_row_fill
            
            current_row += 1
        
        # ============================================
        # FOOTER
        # ============================================
        current_row += 1
        ws.merge_cells(f'A{current_row}:M{current_row}')
        cell = ws[f'A{current_row}']
        cell.value = "END OF REPORT"
        cell.font = Font(bold=True, size=12)
        cell.alignment = center_align
        current_row += 1
        
        ws[f'A{current_row}'] = "Total Records:"
        ws[f'A{current_row}'].font = Font(bold=True)
        ws[f'B{current_row}'] = len(sales_data)
        current_row += 2
        
        ws[f'A{current_row}'] = "Note: All amounts are in Philippine Peso (₱)"
        current_row += 1
        ws[f'A{current_row}'] = "Commission rate: 5% of total order amount"
        current_row += 1
        ws[f'A{current_row}'] = "This is a system-generated report"
        
        # Adjust column widths
        column_widths = {
            'A': 15, 'B': 12, 'C': 12, 'D': 25, 'E': 20,
            'F': 8, 'G': 12, 'H': 12, 'I': 12, 'J': 12,
            'K': 14, 'L': 15, 'M': 40
        }
        
        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width
        
        # Save to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Create response
        response = make_response(output.getvalue())
        response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        response.headers['Content-Disposition'] = f'attachment; filename=sales_report_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        
        return response
        
    except Exception as e:
        print(f"Error exporting Excel: {e}")
        import traceback
        traceback.print_exc()
        return f"Error exporting Excel: {str(e)}", 500
    finally:
        close_db_connection(connection)
